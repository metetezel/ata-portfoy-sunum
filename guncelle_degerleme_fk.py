# Degerleme_CDS_MSCI sekmesindeki "Turkiye F/K", "GOP F/K", "Dunya F/K" ve
# "Turkiye CDS" satırlarını otomatik günceller.
#
# "Turkiye CDS" (DEĞİŞTİ 2026-08-31 — eskiden Mete'nin BBG_Weekly.xlsx'ini
# elle yenilemesine bağlıydı, bkz. Pazartesi_Rutini.md madde 10, ve 07-27'den
# beri hiç dokunulmamış/bayat kalmıştı): artık build_endeks_serisi.py'nin
# ANZ sayfasının CDS mini-grafiği için zaten kullandığı aynı ücretsiz,
# günlük worldgovernmentbonds.com kaynağından (turkiye_cds_serisi())
# besleniyor — ayrı bir manuel adım gerekmiyor. Not: bu satır şu an web
# deck'inin hiçbir sayfasında fiilen gösterilmiyor (Sayfa 24 sadece F/K
# üçlüsünü kullanıyor) — sadece kaynak sekmeyi güncel tutmak için yapılıyor,
# Mete'nin 2026-08-31 tarihli açık isteği üzerine.
#
# Turkiye F/K kaynağı (DEĞİŞTİ 2026-08-17 — eskiden "BIST-100 F/K" adıyla
# CEIC'ten geliyordu, bkz. Pazartesi_Rutini.md madde 20): Sayfa 24'ün
# "Dünya/GOP'a göre iskonto" cümlesi CEIC'in BIST-100 F/K'sını (17,44) MSCI'nin
# GOP F/K'sıyla (17,72) kıyaslıyordu — iki FARKLI sağlayıcı/metodoloji, aradaki
# bilinen ~%7-8 sapma toleransından küçük bir fark (%2) üretiyordu, yani
# istatistiksel olarak anlamsızdı. Mete'nin kararıyla (2026-08-17) CEIC yerine
# MSCI'nin KENDİ Türkiye ülke endeksi (msci-turkey-index-usd-net.pdf) —
# GOP/Dünya F/K'nın geldiği AYNI World factsheet metodolojisi, elma-elma kıyas.
# ⚠️ Kapsam farkı bilerek kabul edildi: bu endeks BIST-100'ün 100 hissesi
# değil, MSCI'nin büyük-sermayeli ~11 şirketlik daha dar Türkiye sepeti (F/K
# tipik olarak BIST-100'den daha düşük çıkar — bu yüzden eski CEIC denemesinde
# "çok uzak" diye reddedilmişti, ama o zamanki hedef "BIST-100'ü temsil eden
# tek bir sayı" bulmaktı; şimdiki hedef "GOP/Dünya ile TUTARLI bir sayı"
# olduğu için tercih tersine döndü). CEIC/Bloomberg kaynaklı eski "BIST-100
# F/K" değeri artık hiçbir yerde kullanılmıyor.
#
# GOP F/K ve Dünya F/K kaynağı:
# https://www.msci.com/documents/10199/255599/msci-world-index.pdf —
# sabit/canlı bir URL, MSCI bu dosyayı her ay sonunda yerinde güncelliyor
# (indirilen PDF'in kendi "generated" tarihi ayın başına denk gelir, içindeki
# veri bir önceki ay sonuna ait — ör. Temmuz başında indirilen dosya "JUN 30"
# verisini taşır). Bu PDF'in "FUNDAMENTALS" tablosu MSCI World'ün YANI SIRA
# karşılaştırma sütunu olarak MSCI Emerging Markets'ı da içeriyor — tek
# indirmeyle her iki rakam da geliyor, ayrı bir EM factsheet'i gerekmiyor.
#
# ⚠️ Bilinen sınırlama (Mete'ye bildirildi, 2026-07-30'da kabul edildi):
#   - MSCI bu factsheet'i sadece AYDA BİR günceller — haftalık deck'te bazı
#     haftalar rakam hiç değişmeyecek, bu normal/beklenen, hata değil.
#   - Rakamlar Bloomberg'in kendi F/K hesaplamasıyla birebir eşleşmeyebilir
#     (farklı endeks kapsamı/metodoloji) — ~%7-8 civarı sapma gözlemlendi,
#     TCMB/Yahoo Finance otomasyonundaki "ondalık bazda birebir değil" kabulüyle
#     aynı kategoride bir tradeoff.
#
# PDF çıkarım tekniği notu: bu PDF'te tablo çizgisi yok (find_tables() 0 tablo
# buluyor), sadece serbest-konumlu metin var. Üstelik "INDEX PERFORMANCE" ve
# "FUNDAMENTALS" tabloları YAN YANA aynı satırları paylaşıyor (ör. "MSCI World"
# etiketi tek başına hem performans hem F/K sütunlarına hizalı) — bu yüzden
# "satırdaki N'inci sayı" gibi düz bir sıralamaya güvenmek yanlış sayıyı
# yakalar (performans tablosunun 8 sayısı önce geliyor). Bunun yerine "P/E"
# sütun başlığının GERÇEK x-koordinatına en yakın sayısal değeri buluyoruz.
import datetime
import os
import re

import fitz
import openpyxl
import requests

from build_endeks_serisi import turkiye_cds_serisi

VERI_KAYNAGI_YOLU = os.path.join(os.path.dirname(__file__), "Veri_Kaynagi.xlsx")
MSCI_WORLD_PDF_URL = "https://www.msci.com/documents/10199/255599/msci-world-index.pdf"
MSCI_TURKEY_PDF_URL = "https://www.msci.com/documents/10199/255599/msci-turkey-index-usd-net.pdf"
TARAYICI_UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/124.0 Safari/537.36"
)

AY_ADLARI_KISA = {
    "JAN": 1, "FEB": 2, "MAR": 3, "APR": 4, "MAY": 5, "JUN": 6,
    "JUL": 7, "AUG": 8, "SEP": 9, "OCT": 10, "NOV": 11, "DEC": 12,
}


def msci_turkiye_fk_oku():
    """MSCI'nin kendi Türkiye ülke endeksinin (MSCI Turkey, USD Net) trailing
    F/K'sını okur — msci_world_em_fk_oku()'nun AYNI koordinat-bazlı çıkarım
    tekniği (tablo çizgisi yok, "P/E" sütun başlığının x-konumuna en yakın
    sayı seçiliyor; "P/E Fwd" ile karışmasın diye trailing "P/E" başlığının
    x'i — solda olan — kullanılıyor, guncelle_peer_fk_msci.py'nin P/E Fwd
    için sağdakini seçmesinin tam tersi). Neden bu endeks: bkz. dosyanın
    başındaki not — GOP/Dünya F/K ile AYNI World factsheet metodolojisi,
    CEIC'in eski "BIST-100 F/K"sından farklı (daha dar, ~11 büyük şirket)
    ama GOP/Dünya ile elma-elma kıyaslanabilir."""
    pdf_bytes = requests.get(MSCI_TURKEY_PDF_URL, headers={"User-Agent": "Mozilla/5.0"}, timeout=20).content
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    page = doc[0]
    kelimeler = page.get_text("words")

    fundamentals_y = next(w[1] for w in kelimeler if w[4] == "FUNDAMENTALS")
    tarih_kelimeleri = [w[4] for w in kelimeler if abs(w[1] - fundamentals_y) <= 1.5 and w[0] > 400]
    ay_kisa = next(t.strip("(") for t in tarih_kelimeleri if t.strip("(").upper() in AY_ADLARI_KISA)
    gun = int(next(t.strip(",") for t in tarih_kelimeleri if t.rstrip(",").isdigit()))
    yil = int(next(t.strip(")") for t in tarih_kelimeleri if t.strip(")").isdigit() and len(t.strip(")")) == 4))
    veri_tarihi = datetime.date(yil, AY_ADLARI_KISA[ay_kisa.upper()], gun)

    pe_header_y = min(w[1] for w in kelimeler if w[4] == "P/E" and w[1] > fundamentals_y)
    pe_x = min(w[0] for w in kelimeler if w[4] == "P/E" and abs(w[1] - pe_header_y) <= 1.5)

    # Bu factsheet'te "MSCI Turkey" satırı FUNDAMENTALS tablosunda birden
    # fazla kez geçebiliyor (sayfanın alt kısmında ilgisiz bir tabloda da
    # var) — ilk eşleşen VE sayısal bir P/E değeri taşıyan satır alınıyor,
    # tıpkı msci_world_em_fk_oku()'daki fk_bul() gibi.
    for aday in (w for w in kelimeler if w[4] == "MSCI" and w[1] > fundamentals_y):
        y = aday[1]
        satir = sorted((w for w in kelimeler if abs(w[1] - y) <= 1.5), key=lambda w: w[0])
        etiket_metni = [w[4] for w in satir if w[0] < 400]
        if etiket_metni[:2] != ["MSCI", "Turkey"]:
            continue
        sayisal = [w for w in satir if re.match(r"^-?\d+\.\d+$", w[4])]
        if not sayisal:
            continue
        en_yakin = min(sayisal, key=lambda w: abs(w[0] - pe_x))
        return veri_tarihi, float(en_yakin[4])

    raise RuntimeError("MSCI Türkiye factsheet'inde FUNDAMENTALS/'MSCI Turkey' satırı bulunamadı — PDF formatı değişmiş olabilir.")


def msci_world_em_fk_oku():
    pdf_bytes = requests.get(MSCI_WORLD_PDF_URL, headers={"User-Agent": "Mozilla/5.0"}, timeout=20).content
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    page = doc[0]
    kelimeler = page.get_text("words")  # (x0,y0,x1,y1,word,block,line,word_no)

    fundamentals_y = next(w[1] for w in kelimeler if w[4] == "FUNDAMENTALS")

    # "FUNDAMENTALS (JUN 30, 2026)" — aynı satırda, sağ tarafta.
    tarih_kelimeleri = [w[4] for w in kelimeler if abs(w[1] - fundamentals_y) <= 1.5 and w[0] > 400]
    ay_kisa = next(t.strip("(") for t in tarih_kelimeleri if t.strip("(").upper() in AY_ADLARI_KISA)
    gun = int(next(t.strip(",") for t in tarih_kelimeleri if t.rstrip(",").isdigit()))
    yil = int(next(t.strip(")") for t in tarih_kelimeleri if t.strip(")").isdigit() and len(t.strip(")")) == 4))
    veri_tarihi = datetime.date(yil, AY_ADLARI_KISA[ay_kisa.upper()], gun)

    # "P/E" sütun başlığının x-konumu — başlık satırında "P/E" kelimesi İKİ
    # kez geçiyor ("P/E" sütununun kendisi ve "P/E Fwd"nin ilk kelimesi olarak)
    # — soldaki (küçük x) gerçek "P/E" sütunudur.
    pe_header_y = min(w[1] for w in kelimeler if w[4] == "P/E" and w[1] > fundamentals_y)
    pe_x = min(w[0] for w in kelimeler if w[4] == "P/E" and abs(w[1] - pe_header_y) <= 1.5)

    def fk_bul(etiket_parcalari):
        ilk = etiket_parcalari[0]
        for aday in (w for w in kelimeler if w[4] == ilk and w[1] > fundamentals_y):
            y = aday[1]
            satir = sorted((w for w in kelimeler if abs(w[1] - y) <= 1.5), key=lambda w: w[0])
            etiket_metni = [w[4] for w in satir if w[0] < 400]
            if etiket_metni[: len(etiket_parcalari)] != etiket_parcalari:
                continue
            sayisal = [w for w in satir if re.match(r"^-?\d+\.\d+$", w[4])]
            if not sayisal:
                continue
            en_yakin = min(sayisal, key=lambda w: abs(w[0] - pe_x))
            return float(en_yakin[4])
        return None

    dunya_fk = fk_bul(["MSCI", "World"])
    gop_fk = fk_bul(["MSCI", "Emerging", "Markets"])

    if dunya_fk is None or gop_fk is None:
        raise RuntimeError(f"FUNDAMENTALS tablosundan P/E okunamadı (Dünya={dunya_fk}, GOP={gop_fk}) — PDF formatı değişmiş olabilir.")

    return veri_tarihi, dunya_fk, gop_fk


def main():
    hedefler = {}  # metrik -> (tarih_str, deger)
    hatalar = []

    try:
        veri_tarihi, dunya_fk, gop_fk = msci_world_em_fk_oku()
        hedefler["Dunya F/K"] = (veri_tarihi.isoformat(), dunya_fk)
        hedefler["GOP F/K"] = (veri_tarihi.isoformat(), gop_fk)
    except Exception as e:
        hatalar.append(f"MSCI World factsheet okuma hatası (GOP/Dünya F/K güncellenmedi): {e}")

    try:
        veri_tarihi, turkiye_fk = msci_turkiye_fk_oku()
        hedefler["Turkiye F/K"] = (veri_tarihi.isoformat(), turkiye_fk)
    except Exception as e:
        hatalar.append(f"MSCI Türkiye factsheet okuma hatası (Turkiye F/K güncellenmedi): {e}")

    try:
        cds_pencere, _cds_istatistik = turkiye_cds_serisi()
        cds_son_tarih = max(cds_pencere)
        hedefler["Turkiye CDS"] = (cds_son_tarih.isoformat(), cds_pencere[cds_son_tarih])
    except Exception as e:
        hatalar.append(f"worldgovernmentbonds.com CDS okuma hatası (Turkiye CDS güncellenmedi): {e}")

    wb = openpyxl.load_workbook(VERI_KAYNAGI_YOLU)
    ws = wb["Degerleme_CDS_MSCI"]

    bulunanlar = set()
    for row in ws.iter_rows(min_row=2):
        metrik = row[1].value
        if metrik in hedefler:
            tarih_str, deger = hedefler[metrik]
            eski = row[2].value
            row[0].value = tarih_str
            row[2].value = round(deger, 2)
            bulunanlar.add(metrik)
            print(f"  {metrik}: {eski} -> {row[2].value} ({tarih_str})")

    # Sekmede satır hiç yoksa (ör. ilk çalıştırma) ekle.
    son_dolu_satir = ws.max_row
    for metrik, (tarih_str, deger) in hedefler.items():
        if metrik in bulunanlar:
            continue
        son_dolu_satir += 1
        ws.cell(row=son_dolu_satir, column=1, value=tarih_str)
        ws.cell(row=son_dolu_satir, column=2, value=metrik)
        ws.cell(row=son_dolu_satir, column=3, value=round(deger, 2))
        print(f"  {metrik}: (yeni satır) -> {round(deger, 2)} ({tarih_str})")

    wb.save(VERI_KAYNAGI_YOLU)
    print("OK — Degerleme_CDS_MSCI güncellendi.")
    if hatalar:
        print("UYARI — bazı kaynaklar okunamadı:")
        for h in hatalar:
            print(f"  {h}")


if __name__ == "__main__":
    main()
