# TEK SEFERLİK göç script'i — Sunum.xlsx, Copy of MCAP_to_GDP.xlsx ve MXWO vs
# MXEF (BB).xlsx'in içindeki DONUK (bir daha hiç güncellenmeyecek) tarihi
# verileri kendi kalıcı JSON anlık görüntülerimize çıkarır, böylece bu 3
# büyük Excel dosyası projeden tamamen silinebilir hale gelir.
#
# NEDEN (2026-07-30, Mete'nin isteğiyle): "ne kadar az dosya olursa projeyi
# o kadar iyi şekilde başka insanlar kullanabilir/geliştirebilir" — bu 3
# dosya zaten Mete tarafından bir daha güncellenmeyecek (build_endeks_serisi.py
# ve build_fiyat_serisi.py'nin hibrit tasarımı zaten bunu varsayıyordu), ama
# script'ler hâlâ her çalıştırmada bu dosyaları CANLI OKUYORDU — yani dosyalar
# hâlâ "silinemez" durumdaydı. Bu script o canlı bağımlılığı koparıyor: aynı
# donuk veriyi bir kere JSON'a döküyor, script'ler artık JSON'dan okuyor.
#
# Bu script HAFTALIK RUTİNİN PARÇASI DEĞİL — bir daha çalıştırmaya gerek yok
# (kaynak dosyalar zaten silinecek). Sadece göç anını belgelemek için saklanıyor.
import datetime
import json
import os

import openpyxl

PROJE_KOK = os.path.dirname(__file__)
CIKTI_KLASORU = os.path.join(PROJE_KOK, "sabit_kaynaklar")
os.makedirs(CIKTI_KLASORU, exist_ok=True)

# ANZ'nin TEFAS'a devrettiği sınır tarihi (build_fiyat_serisi.py'deki
# "tefas_baslangic" ile birebir aynı) — bu tarihten SONRASI zaten TEFAS'tan
# geliyor, Sunum.xlsx'ten sadece bu tarihten ÖNCESİNİ dondurmamız yeterli.
ANZ_TEFAS_SINIRI = datetime.date(2021, 8, 1)


def anz_eurobond_dondur():
    yol = os.path.join(PROJE_KOK, "Sunum.xlsx")
    wb = openpyxl.load_workbook(yol, data_only=True, read_only=True)
    ws = wb["Eurobond"]
    seri = {}
    for row in ws.iter_rows(values_only=True):
        tarih, deger = row[0], row[7]
        if not isinstance(tarih, datetime.datetime) or not isinstance(deger, (int, float)):
            continue
        d = tarih.date()
        if d >= ANZ_TEFAS_SINIRI:
            continue  # bu noktadan sonrası zaten TEFAS'tan geliyor
        seri[d.isoformat()] = float(deger)
    wb.close()

    cikti_yolu = os.path.join(CIKTI_KLASORU, "anz_eurobond_gecmis.json")
    with open(cikti_yolu, "w", encoding="utf-8") as f:
        json.dump(seri, f, ensure_ascii=False)
    print(f"OK — {cikti_yolu} ({len(seri)} gün, {min(seri)} - {max(seri)})")


def mcap_gdp_dondur():
    yol = os.path.join(PROJE_KOK, "Copy of MCAP_to_GDP.xlsx")
    wb = openpyxl.load_workbook(yol, data_only=True)
    ws = wb["MCAP to GDP 93 (xu100)"]
    seri = {}
    istatistik = None
    for row in ws.iter_rows(min_row=3, values_only=True):
        tarih, oran = row[0], row[3]
        if not isinstance(tarih, datetime.datetime) or not isinstance(oran, (int, float)):
            continue
        d = tarih.date()
        if d > datetime.date.today():
            continue
        seri[d.isoformat()] = float(oran)
        if istatistik is None and isinstance(row[4], (int, float)):
            istatistik = {
                "ortalama": round(row[4], 4),
                "ustSapma": round(row[5], 4),
                "altSapma": round(row[6], 4),
                "maksimum": round(row[7], 4),
                "minimum": round(row[8], 4),
            }
    wb.close()

    cikti_yolu = os.path.join(CIKTI_KLASORU, "mcap_gdp_gecmis.json")
    with open(cikti_yolu, "w", encoding="utf-8") as f:
        json.dump({"istatistik": istatistik, "seri": seri}, f, ensure_ascii=False)
    print(f"OK — {cikti_yolu} ({len(seri)} gün, {min(seri)} - {max(seri)})")
    print(f"  istatistik: {istatistik}")


def mxwo_mxef_dondur():
    yol = os.path.join(PROJE_KOK, "MXWO vs MXEF (BB).xlsx")
    wb = openpyxl.load_workbook(yol, data_only=True)
    ws = wb["vs"]
    seri = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        tarih, mxwo, mxef = row[2], row[3], row[4]
        if not isinstance(tarih, datetime.datetime):
            continue
        if not isinstance(mxwo, (int, float)) or not isinstance(mxef, (int, float)):
            continue
        if mxwo <= 0 or mxef <= 0:
            continue
        d = tarih.date()
        if d > datetime.date.today():
            continue
        seri[d.isoformat()] = [float(mxwo), float(mxef)]
    wb.close()

    cikti_yolu = os.path.join(CIKTI_KLASORU, "mxwo_mxef_gecmis.json")
    with open(cikti_yolu, "w", encoding="utf-8") as f:
        json.dump(seri, f, ensure_ascii=False)
    print(f"OK — {cikti_yolu} ({len(seri)} gün, {min(seri)} - {max(seri)})")


if __name__ == "__main__":
    anz_eurobond_dondur()
    mcap_gdp_dondur()
    mxwo_mxef_dondur()
    print()
    print("Tamamlandı. build_fiyat_serisi.py ve build_endeks_serisi.py artık bu")
    print("JSON'ları okuyacak şekilde güncellenmeli — bu script'in kendisi onları")
    print("değiştirmiyor, sadece veriyi donduruyor.")
