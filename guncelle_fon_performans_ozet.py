# Refreshes both "Ata Fonları: Güçlü Performans" pages' data from the same
# live COPY Fon Broşür.xlsx used by guncelle_fon_getirileri.py — confirmed to
# be the right source by finding the legacy Sunum.xlsx's own "DİJİTAL" sheet,
# which pulls these exact pages' numbers from 'ATA FON PERFORMANS (Kısa)' via
# external-link formulas (2026-07-28).
#
# The legacy deck has TWO of these chart pages, not one: a trailing-12-month
# version (page "03") and a Yılbaşından-Beri version (page "04") — same idea,
# same source, but a genuinely different fund/benchmark selection and
# grouping in each (not just the same list re-sorted), so each gets its own
# row-definition list below rather than trying to derive one from the other.
#
# Both target sheets are fully automated — there's no manual-only subset to
# preserve here, so every run fully replaces each sheet's content.
import os

import openpyxl

from fon_veri_ortak import (
    kisa_sayfasini_ayristir,
    piyasa_benchmarklarini_oku,
    rapor_tarihini_oku,
    temettu_dahil_yillik_oku,
)

FON_BROSUR_YOLU = os.environ.get(
    "FON_BROSUR_YOLU",
    r"\\atafiles\Ata.Portföy\Farshad\Fonlar\COPY Fon Broşür.xlsx",
)
VERI_KAYNAGI_YOLU = os.path.join(os.path.dirname(__file__), "Veri_Kaynagi.xlsx")

# (Grup, Ad, Tip, kaynak) — kaynak is one of:
#   {"kod": "AAV"}                 -> that fund's own value for this page's pencere
#   {"temettu_kod": "AYA"}         -> that fund's dividend-inclusive Yillik row (Yillik page only)
#   {"standalone": "..."}          -> a name from the standalone benchmark table
YILLIK_TANIMLARI = [
    (1, "Ata İkinci Hisse Senedi Fonu", "Fon", {"kod": "AAV"}),
    (1, "Ata Kar Payı Ödeyen Hisse Senedi Fonu", "Fon", {"kod": "AYA"}),
    (1, "Ata Kar Payı Ödeyen Hisse Senedi Fonu (Temettü Dahil)", "Fon", {"temettu_kod": "AYA"}),
    (1, "Ata Katılım Hisse Senedi Fonu", "Fon", {"kod": "TLZ"}),
    (1, "BIST 100 Getiri Endeksi", "Benchmark", {"standalone": "BIST 100 Getiri Endeksi"}),
    (1, "BIST Temettü 25 Getiri Endeksi", "Benchmark", {"standalone": "BIST Temettü 25 Getiri Endeksi"}),
    (2, "Ata Birinci Değişken Fon", "Fon", {"kod": "AED"}),
    (3, "Ata Para Piyasası Serbest (TL) Fon", "Fon", {"kod": "DGH"}),
    (3, "Ata Para Piyasası Fonu", "Fon", {"kod": "AAL"}),
    (3, "KYD Mevduat Endeksi", "Benchmark", {"standalone": "KYD Mevduat Endeksi"}),
    (4, "Ata Tarım ve Gıda Değişken Fon", "Fon", {"kod": "YLC"}),
    (4, "Ata Robotik Teknolojileri Değişken Fon", "Fon", {"kod": "RTG"}),
    (4, "Ata Havacılık ve Savunma Teknolojileri Değişken Fon", "Fon", {"kod": "JET"}),
    (4, "Ata Enerji Değişken Fon", "Fon", {"kod": "URA"}),
    (5, "Ata Altın Katılım Fonu", "Fon", {"kod": "PKF"}),
    (5, "Ata Fon Sepeti Serbest Fon", "Fon", {"kod": "AAS"}),
    (5, "Ata Dördüncü Serbest (Eurobond) Fon (TL Getiri)", "Fon", {"kod": "ANZ"}),
    (5, "Ata Dördüncü Serbest (Eurobond) Fon (USD Getiri)", "Fon", {"kod": "UANZ"}),
    (5, "USD/TL", "Benchmark", {"standalone": "USD/TL"}),
    (5, "TÜFE ({yil} - Yıllık)", "Benchmark", {"standalone": "TÜFE"}),
]

YBB_TANIMLARI = [
    (1, "İkinci Hisse Senedi Fonu", "Fon", {"kod": "AAV"}),
    (1, "Kar Payı Ödeyen Hisse Senedi Fonu", "Fon", {"kod": "AYA"}),
    (1, "Katılım Hisse Senedi (TL) Fonu", "Fon", {"kod": "TLZ"}),
    (2, "Enerji Değişken Fon", "Fon", {"kod": "URA"}),
    (2, "Havacılık ve Savunma Teknolojileri Fonu", "Fon", {"kod": "JET"}),
    (2, "Tarım ve Gıda Değişken Fon", "Fon", {"kod": "YLC"}),
    (2, "Robotik Teknolojileri Değişken Fon", "Fon", {"kod": "RTG"}),
    (2, "BIST 100 Getiri Endeksi", "Benchmark", {"standalone": "BIST 100 Getiri Endeksi"}),
    (2, "BIST Temettü 25 Getiri Endeksi", "Benchmark", {"standalone": "BIST Temettü 25 Getiri Endeksi"}),
    (3, "ATA Birinci Değişken Fon", "Fon", {"kod": "AED"}),
    (3, "KYD Mevduat Endeksi", "Benchmark", {"standalone": "KYD Mevduat Endeksi"}),
    (4, "Ata Altın Katılım Fonu", "Fon", {"kod": "PKF"}),
    (4, "ATA 4. Serbest (EB) Fon (TL Getiri)", "Fon", {"kod": "ANZ"}),
    (4, "ATA 4. Serbest (EB) Fon (USD Getiri)", "Fon", {"kod": "UANZ"}),
    (4, "ATA Fon Sepeti Serbest Fon", "Fon", {"kod": "AAS"}),
    (4, "USD/TRY", "Benchmark", {"standalone": "USD/TL"}),
    (5, "Ata Para Piyasası Serbest (TL) Fon", "Fon", {"kod": "DGH"}),
    (5, "Ata Para Piyasası Fonu", "Fon", {"kod": "AAL"}),
    (5, "TÜFE ({yil} - Yılbaşından Beri)", "Benchmark", {"standalone": "TÜFE"}),
]

FON_KODLARI = {"AAV", "AYA", "TLZ", "AED", "DGH", "AAL", "YLC", "RTG", "JET", "URA", "PKF", "AAS", "ANZ", "UANZ"}


def guncelle(hedef_wb, sayfa_adi, tanimlar, pencere, fon_verileri, aya_temettu_dahil, standalone_verileri, rapor_tarihi):
    satirlar = []
    eksikler = []
    for grup, ad_sablonu, tip, kaynak in tanimlar:
        ad = ad_sablonu.format(yil=rapor_tarihi.year if rapor_tarihi else "")
        if "kod" in kaynak:
            deger = fon_verileri.get(kaynak["kod"], {}).get(pencere, (None, None))[0]
        elif "temettu_kod" in kaynak:
            deger = aya_temettu_dahil
        else:
            deger = standalone_verileri.get(kaynak["standalone"])
        if deger is None:
            eksikler.append(ad)
            continue
        satirlar.append((grup, ad, deger, tip))

    if eksikler:
        print(f"UYARI ({sayfa_adi}): değer bulunamayan satırlar (atlandı): {eksikler}")

    ws = hedef_wb[sayfa_adi]
    ws.delete_rows(2, ws.max_row)
    for r, row_data in enumerate(satirlar, start=2):
        for c, val in enumerate(row_data, start=1):
            ws.cell(row=r, column=c, value=val)
    print(f"OK ({sayfa_adi}) — {len(satirlar)}/{len(tanimlar)} satır yazıldı.")


def main():
    kaynak_wb = openpyxl.load_workbook(FON_BROSUR_YOLU, data_only=True, read_only=True)
    kisa_ws = kaynak_wb["ATA FON PERFORMANS (Kısa)"]
    rapor_tarihi = rapor_tarihini_oku(kisa_ws)
    fon_verileri = kisa_sayfasini_ayristir(kisa_ws, FON_KODLARI)
    aya_temettu_dahil = temettu_dahil_yillik_oku(kisa_ws, "AYA")
    standalone_yillik = piyasa_benchmarklarini_oku(kisa_ws, "Yillik")
    standalone_ybb = piyasa_benchmarklarini_oku(kisa_ws, "Yilbasindan Beri")
    kaynak_wb.close()

    hedef_wb = openpyxl.load_workbook(VERI_KAYNAGI_YOLU)
    guncelle(hedef_wb, "Ata_Fonlari_Performans_Ozet", YILLIK_TANIMLARI, "Yillik",
             fon_verileri, aya_temettu_dahil, standalone_yillik, rapor_tarihi)
    guncelle(hedef_wb, "Ata_Fonlari_Performans_Ozet_YBB", YBB_TANIMLARI, "Yilbasindan Beri",
             fon_verileri, aya_temettu_dahil, standalone_ybb, rapor_tarihi)
    hedef_wb.save(VERI_KAYNAGI_YOLU)
    print(f"Rapor tarihi: {rapor_tarihi}")


if __name__ == "__main__":
    main()
