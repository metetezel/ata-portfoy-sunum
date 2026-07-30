# Refreshes Fon_Performans's short-term windows (Aylik/3 Aylik/6 Aylik/Yillik/
# Yilbasindan Beri) in Veri_Kaynagi.xlsx directly from the live Rasyonet/EquityRT
# -fed fund engine, instead of Mete hand-typing them from a rendered PDF each
# week.
#
# Originally scoped narrow (per Mete, 2026-07-28): only the five short-term
# windows automated, since 3 Yil/5 Yil/Kurulustan Beri aren't precomputed
# anywhere in the source workbook and would need computing from the raw
# Price/Bench daily series instead — a separate, bigger task. Extended
# 2026-07-29 for AYA's Kurulustan Beri specifically (see OTOMATIK_EK_PENCERELER
# below and fon_veri_ortak.kurulustan_beri_hesapla) — other funds' equivalent
# rows (e.g. AAV's 3 Yil/5 Yil/Kurulustan Beri) stay manual/preserved until
# asked for the same treatment.
#
# Run this whenever fresh numbers are wanted (it's meant to run daily — the
# source file itself is already refreshed every morning by an existing human
# process, so there's no "trigger a refresh" step here, just "read what's
# already fresh"). Re-running build_veri_kaynagi.py from scratch would wipe
# this — that script is for schema changes only, this one is for data refresh,
# and they intentionally don't overlap in responsibility.
import datetime
import os

import openpyxl

from fon_veri_ortak import (
    PENCERE_KOLONU,
    formul_katilim100_getirisi_oku,
    gunluk_seriden_pencereler_hesapla,
    kisa_sayfasini_ayristir,
    kurulustan_beri_hesapla,
    piyasa_benchmarklarini_tum_pencereler_oku,
    rapor_tarihini_oku,
)

FON_BROSUR_YOLU = os.environ.get(
    "FON_BROSUR_YOLU",
    r"\\atafiles\Ata.Portföy\Farshad\Fonlar\COPY Fon Broşür.xlsx",
)
VERI_KAYNAGI_YOLU = os.path.join(os.path.dirname(__file__), "Veri_Kaynagi.xlsx")

# Fon_Katalog's 14 codes — only these get automated rows; anything else in the
# source sheet (other funds Ata runs that don't have their own deck page) is
# ignored here.
FON_KODLARI = {"AYA", "AAV", "TLZ", "URA", "JET", "YLC", "RTG", "AED", "AAS", "ANZ", "DGH", "AAL", "PKF", "PKP"}


def main():
    kaynak_wb = openpyxl.load_workbook(FON_BROSUR_YOLU, data_only=True, read_only=True)
    kisa_ws = kaynak_wb["ATA FON PERFORMANS (Kısa)"]
    rapor_tarihi = rapor_tarihini_oku(kisa_ws)
    veriler = kisa_sayfasini_ayristir(kisa_ws, FON_KODLARI)

    # Confirmed real bug in the source sheet (2026-07-29): AYA's own adjacent
    # "Benchmark" row is mislabeled "Benchmark-BIST 100" and literally holds
    # BIST 100 Getiri Endeksi's figures (byte-for-byte identical to the
    # standalone BIST 100 row) — not AYA's real benchmark, BIST Temettü 25
    # Getiri Endeksi (Fon_Katalog.Benchmark_Aciklama), which lives instead in
    # the sheet's separate "Piyasa Benchmarkları" section. Same class of
    # stale-reference bug as the legacy DİJİTAL sheet's row-drift issue —
    # override with the correct standalone series rather than trusting the
    # adjacent row for this one fund.
    standalone_tum = piyasa_benchmarklarini_tum_pencereler_oku(kisa_ws)
    temettu_25 = standalone_tum.get("BIST Temettü 25 Getiri Endeksi")
    if "AYA" in veriler and temettu_25:
        for pencere, dogru_bench_v in temettu_25.items():
            if pencere in veriler["AYA"]:
                fon_v, _yanlis_bench_v = veriler["AYA"][pencere]
                veriler["AYA"][pencere] = (fon_v, dogru_bench_v)

    # Same bug, different fund (2026-07-29): TLZ's real benchmark (BIST
    # Katılım 100 Getiri / XK100) isn't in the standalone Piyasa
    # Benchmarkları table at all — its correct return figures live in the
    # separate "Formül" sheet instead (see formul_katilim100_getirisi_oku).
    formul_ws = kaynak_wb["Formül"]
    katilim100 = formul_katilim100_getirisi_oku(formul_ws)
    if "TLZ" in veriler and katilim100:
        for pencere, dogru_bench_v in katilim100.items():
            if pencere in veriler["TLZ"]:
                fon_v, _yanlis_bench_v = veriler["TLZ"][pencere]
                veriler["TLZ"][pencere] = (fon_v, dogru_bench_v)

    # Third instance of this same general class of bug (2026-07-29): ANZ's
    # adjacent "Benchmark***" row isn't mislabeled data exactly, but it's a
    # genuinely DIFFERENT real benchmark than Fon_Katalog states and the
    # legacy deck displays — footnoted as "ANZ Benchmark: KYD 1 Aylık USD
    # Mevduat (TL)" (TL-denominated, so it bakes in USD/TL currency
    # appreciation on top of the deposit rate — much bigger number) — not
    # the plain USD-denominated "KYD USD Mevduat" the fund is actually
    # benchmarked against everywhere else in this project.
    kyd_usd_mevduat = standalone_tum.get("KYD USD Mevduat")
    if "ANZ" in veriler and kyd_usd_mevduat:
        for pencere, dogru_bench_v in kyd_usd_mevduat.items():
            if pencere in veriler["ANZ"]:
                fon_v, _yanlis_bench_v = veriler["ANZ"][pencere]
                veriler["ANZ"][pencere] = (fon_v, dogru_bench_v)

    # Fourth instance of this class of bug, found on AED (2026-07-29): its
    # adjacent "Benchmark"/"Değişken Fon" row is a weighted composite (15%
    # KYD Repo + 55% BIST100 Getiri + 15% KYD Altın + 15% KYD Eurobond TL,
    # confirmed via the Formül sheet's own weight table) — a real number,
    # but not what the legacy deck displays. AED's page actually shows
    # THREE separate raw comparisons (BIST-100 Getiri / USD / TL Mevduat),
    # not one blended figure — the primary Benchmark_Getiri_Yuzde column
    # here becomes BIST-100 Getiri (matching Fon_Katalog.Benchmark_Kisa_Ad);
    # the other two ("USD" = USD/TL, "TL Mevduat" = KYD Mevduat Endeksi) go
    # to Fon_Performans_Ek_Kriter below instead, since the normal single-
    # benchmark shape has no room for a third column.
    bist100_getiri = standalone_tum.get("BIST 100 Getiri Endeksi")
    if "AED" in veriler and bist100_getiri:
        for pencere, dogru_bench_v in bist100_getiri.items():
            if pencere in veriler["AED"]:
                fon_v, _yanlis_bench_v = veriler["AED"][pencere]
                veriler["AED"][pencere] = (fon_v, dogru_bench_v)

    # Fifth instance, different shape (2026-07-29): DGH has NO adjacent
    # "Benchmark" row at all in this sheet (kisa_sayfasini_ayristir finds
    # nothing to pair with its Fon row) — its real benchmark, "KYD Mevduat
    # Endeksi" (matches Fon_Katalog and the legacy deck's "KYD Mevduat(2)"
    # column almost exactly), only lives in the standalone table.
    kyd_mevduat = standalone_tum.get("KYD Mevduat Endeksi")
    if "DGH" in veriler and kyd_mevduat:
        for pencere, dogru_bench_v in kyd_mevduat.items():
            if pencere in veriler["DGH"]:
                fon_v, _bos = veriler["DGH"][pencere]
                veriler["DGH"][pencere] = (fon_v, dogru_bench_v)

    # DGH's third column, "KYD ÖST Değişken" — not precomputed anywhere,
    # only exists as a raw daily series (Bench sheet, unlabeled column 14,
    # found via the row-3782 stray-label search: 'ÖzelSektör endeksi
    # değişken'). Computed here (not in build_fiyat_serisi.py) since it's
    # needed as return-table percentages, not a price-chart line.
    bench_ws = kaynak_wb["Bench"]
    dgh_ost_degisken = gunluk_seriden_pencereler_hesapla(bench_ws, 14)

    # AYA's "Kuruluştan Beri" row — automated 2026-07-29 (Mete's ask), no
    # longer a one-off manual entry frozen since whenever it was last typed.
    # Same raw Price/Bench columns already validated in build_fiyat_serisi.py's
    # FON_KAYNAK["AYA"] (Price col 7, fund rebased from its own 2010-06-17
    # start; Bench col 38, benchmark rebased from ITS OWN start since BIST
    # Temettü 25 didn't exist before 2011-07-01) — reproduces the previous
    # manual figures (fund +4554.8%, benchmark +4994.2%) exactly.
    price_ws = kaynak_wb["Price"]
    aya_fon_kurulus = kurulustan_beri_hesapla(price_ws, 7, baslangic_tarihi=datetime.date(2010, 6, 17))
    aya_bench_kurulus = kurulustan_beri_hesapla(bench_ws, 38)
    if "AYA" in veriler and aya_fon_kurulus is not None:
        veriler["AYA"]["Kurulustan Beri"] = (aya_fon_kurulus, aya_bench_kurulus)

    # AAV's "Kuruluştan Beri" — same treatment, Mete asked for it 2026-07-29
    # alongside AAV's 3 Yıl/5 Yıl. Fund: Price
    # col 14 ("AAV:TMF", FON_KAYNAK["AAV"]), rebased from its own 2012-11-16
    # start. Benchmark: Bench col 24 (BIST100 Getiri) — UNLIKE AYA, this
    # benchmark already existed well before AAV's own start (its own earliest
    # data is 2009), so it's rebased from AAV's start date too, not its own
    # earliest date (matching karsilastirma_hazirla's "index doesn't exist
    # yet" exception being the ONLY reason to use a benchmark's own start).
    # Reproduces the previous manual figures (fund +5924.4%, benchmark
    # +2694.2%) within ~0.6% — the same small, already-accepted drift this
    # exact Price/Bench pairing showed for AAV's Aylık/Yıllık figures earlier
    # this session (AAV's own data is a bit noisier than AYA's, not a sign of
    # a wrong source).
    aav_fon_kurulus = kurulustan_beri_hesapla(price_ws, 14, baslangic_tarihi=datetime.date(2012, 11, 16))
    aav_bench_kurulus = kurulustan_beri_hesapla(bench_ws, 24, baslangic_tarihi=datetime.date(2012, 11, 16))
    if "AAV" in veriler and aav_fon_kurulus is not None:
        veriler["AAV"]["Kurulustan Beri"] = (aav_fon_kurulus, aav_bench_kurulus)

    # AAV's "3 Yıl"/"5 Yıl" — ALSO automated (Mete's explicit call, 2026-07-29,
    # after being told the raw-price-ratio approach doesn't closely match the
    # current manual figures: computed fund 3 Yıl +131.5%/5 Yıl +1233.3% vs
    # the old manual +157.9%/+1298.3%, a 5-17% relative gap, much bigger than
    # the <1% drift accepted everywhere else in this project). Checked for a
    # precomputed alternative first (the "Yıllık Compare" sheet, by name the
    # most likely candidate) but it's a tiny unlabeled 6-row scratch area,
    # not a real source — no better option was found, so this uses the same
    # raw-ratio computation despite the gap, per Mete's decision to prefer a
    # live (if imperfect) figure over a frozen manual one.
    aav_fon_3yil = kurulustan_beri_hesapla(price_ws, 14, yil_sayisi=3)
    aav_bench_3yil = kurulustan_beri_hesapla(bench_ws, 24, yil_sayisi=3)
    aav_fon_5yil = kurulustan_beri_hesapla(price_ws, 14, yil_sayisi=5)
    aav_bench_5yil = kurulustan_beri_hesapla(bench_ws, 24, yil_sayisi=5)
    if "AAV" in veriler:
        if aav_fon_3yil is not None:
            veriler["AAV"]["3 Yil"] = (aav_fon_3yil, aav_bench_3yil)
        if aav_fon_5yil is not None:
            veriler["AAV"]["5 Yil"] = (aav_fon_5yil, aav_bench_5yil)

    # USD/TRY's own 3 Yıl/5 Yıl change — read here, BEFORE kaynak_wb closes,
    # for Fon_Yillik_Getiri_Karsilastirma's USD rows further below (that
    # block runs after kaynak_wb.close(), so it can't read from bench_ws
    # directly; these two plain floats are all it needs). Column 9, labeled
    # "DOLAR" via the same row-3782/row-3528 stray-label search used
    # throughout this project — a genuine raw USD/TRY series back to 1997.
    usd_3yil_degisim = kurulustan_beri_hesapla(bench_ws, 9, yil_sayisi=3)
    usd_5yil_degisim = kurulustan_beri_hesapla(bench_ws, 9, yil_sayisi=5)

    kaynak_wb.close()

    eksik = FON_KODLARI - veriler.keys()
    if eksik:
        print(f"UYARI: kaynak sayfada bulunamayan fon kodları: {sorted(eksik)}")

    hedef_wb = openpyxl.load_workbook(VERI_KAYNAGI_YOLU)
    ws = hedef_wb["Fon_Performans"]

    mevcut_satirlar = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Any real data row always has a Fon_Kodu — this also catches the
        # single-column "↑ örnek satır — silip kendi verinizi girin" note
        # row (row[0] set, row[1] None), which the old "both None" check
        # missed and left as a permanent stray row surviving every refresh.
        if row[1] is None:
            continue
        mevcut_satirlar.append(row)

    otomatik_pencereler = set(PENCERE_KOLONU.keys())
    # Per-fund windows automated beyond the standard 5 above — grows as more
    # funds' 3 Yıl/5 Yıl/Kuruluştan Beri rows get computed live instead of
    # staying one-off manual entries. AYA's "Kurulustan Beri" was the first
    # (2026-07-29); other funds' equivalent rows (e.g. AAV's) are still
    # manual and stay preserved until asked for the same treatment.
    OTOMATIK_EK_PENCERELER = {"AYA": {"Kurulustan Beri"}, "AAV": {"Kurulustan Beri", "3 Yil", "5 Yil"}}
    korunan_satirlar = [
        r for r in mevcut_satirlar
        if not (
            r[1] in FON_KODLARI
            and (r[2] in otomatik_pencereler or r[2] in OTOMATIK_EK_PENCERELER.get(r[1], set()))
        )
    ]

    tarih_str = rapor_tarihi.isoformat() if rapor_tarihi else mevcut_satirlar[0][0] if mevcut_satirlar else None
    yeni_satirlar = []
    for fon_kodu, pencereler in veriler.items():
        for pencere, (fon_v, bench_v) in pencereler.items():
            if fon_v is None:
                continue  # kaynakta #REF!/#DIV/0! vb. hata varsa bu satırı atla
            yeni_satirlar.append((tarih_str, fon_kodu, pencere, fon_v, bench_v))

    ws.delete_rows(2, ws.max_row)
    for r, row_data in enumerate(korunan_satirlar + yeni_satirlar, start=2):
        for c, val in enumerate(row_data, start=1):
            ws.cell(row=r, column=c, value=val)

    # Fon_Getiri_Analizi_Detay (AAV's Sayfa 10, Net/Brüt x Fon/Benchmark/
    # Nispi detail table) — automated 2026-07-29 alongside the row above.
    # Reason: automating ONLY Fon_Performans's Kurulustan Beri (above) made
    # this page disagree with page 9 by ~0.6% (two adjacent AAV pages
    # showing different numbers for "the same fact") — Mete's call: keep
    # both in sync, automate this page too, even though it means also
    # picking up the 3 Yıl/5 Yıl gap flagged above.
    #
    # "Net" side is NOTHING but this table's own Yillik/3 Yil/5 Yil/
    # Kurulustan Beri figures — this whole page's "Net" convention already
    # IS what Fon_Performans represents for every fund, so these are reused
    # directly, not recomputed independently. "Yıllıklandırılmış" columns
    # are the standard CAGR annualization of the cumulative ones (validated:
    # this exact formula reproduces the OLD manual Brüt figures' own
    # Yıllıklandırılmış columns from their own Kümülatif columns exactly, so
    # it's definitely the right formula, whatever computed the Kümülatif
    # figures themselves). "Nispi" (relative) = Fon − Benchmark, simple
    # subtraction (validated against the old data: e.g. Net Yıllık Nispi
    # -15.3 = 18.3-33.6 exactly) — not a ratio/alpha calc.
    #
    # "Brüt" (gross of management fee) has NO live source anywhere in this
    # project and reconstructing Rasyonet's exact fee-accrual methodology
    # isn't documented — a flat 2.25% (AAV's stated Yönetim Ücreti) grossed
    # up over each window's years came close to the old Brüt figures but not
    # exactly (drifting further on longer windows), so instead of guessing a
    # methodology, this CARRIES FORWARD the exact Brüt/Net ratio the current
    # manual figures already imply, per window, onto the new Net numbers —
    # keeps Brüt internally consistent with Net without inventing a new,
    # unvalidated formula. Benchmark has no Net/Brüt distinction (an index
    # has no management fee) — Brüt/Benchmark is identical to Net/Benchmark,
    # matching the old data exactly.
    if "AAV" in veriler and aav_fon_kurulus is not None:
        aav_p = veriler["AAV"]
        yillik_v, bench_yillik_v = aav_p["Yillik"]
        uc_yil_v, bench_uc_yil_v = aav_p.get("3 Yil", (None, None))
        bes_yil_v, bench_bes_yil_v = aav_p.get("5 Yil", (None, None))
        kurulus_v, bench_kurulus_v = aav_p["Kurulustan Beri"]

        def yillandir(kumulatif_yuzde, yil_sayisi):
            return round(((1 + kumulatif_yuzde / 100) ** (1 / yil_sayisi) - 1) * 100, 1)

        aav_kurulus_tarihi = datetime.date(2012, 11, 16)
        kurulus_yil_sayisi = ((rapor_tarihi or datetime.date.today()) - aav_kurulus_tarihi).days / 365.25

        net_fon = {
            "Yillik": yillik_v, "3Yil_Kumulatif": uc_yil_v, "5Yil_Kumulatif": bes_yil_v,
            "BaslangictanBeri_Kumulatif": kurulus_v,
            "3Yil_Yillik": yillandir(uc_yil_v, 3) if uc_yil_v is not None else None,
            "5Yil_Yillik": yillandir(bes_yil_v, 5) if bes_yil_v is not None else None,
            "BaslangictanBeri_Yillik": yillandir(kurulus_v, kurulus_yil_sayisi),
        }
        net_bench = {
            "Yillik": bench_yillik_v, "3Yil_Kumulatif": bench_uc_yil_v, "5Yil_Kumulatif": bench_bes_yil_v,
            "BaslangictanBeri_Kumulatif": bench_kurulus_v,
            "3Yil_Yillik": yillandir(bench_uc_yil_v, 3) if bench_uc_yil_v is not None else None,
            "5Yil_Yillik": yillandir(bench_bes_yil_v, 5) if bench_bes_yil_v is not None else None,
            "BaslangictanBeri_Yillik": yillandir(bench_kurulus_v, kurulus_yil_sayisi),
        }
        net_nispi = {k: round(net_fon[k] - net_bench[k], 1) for k in net_fon}

        # Empirically-implied Brüt/Net ratio per cumulative column, solved
        # from the CURRENT (about-to-be-overwritten) manual figures — see
        # comment above for why this carries the ratio forward instead of
        # assuming a fee rate.
        ESKI_NET_FON = {"Yillik": 18.3, "3Yil_Kumulatif": 157.9, "5Yil_Kumulatif": 1298.3, "BaslangictanBeri_Kumulatif": 5924.4}
        ESKI_BRUT_FON = {"Yillik": 20.9, "3Yil_Kumulatif": 172.5, "5Yil_Kumulatif": 1433.6, "BaslangictanBeri_Kumulatif": 7553.3}
        tasima_orani = {k: (1 + ESKI_BRUT_FON[k] / 100) / (1 + ESKI_NET_FON[k] / 100) for k in ESKI_NET_FON}

        brut_fon = {k: round(((1 + net_fon[k] / 100) * tasima_orani[k] - 1) * 100, 1) for k in tasima_orani}
        brut_fon["3Yil_Yillik"] = yillandir(brut_fon["3Yil_Kumulatif"], 3)
        brut_fon["5Yil_Yillik"] = yillandir(brut_fon["5Yil_Kumulatif"], 5)
        brut_fon["BaslangictanBeri_Yillik"] = yillandir(brut_fon["BaslangictanBeri_Kumulatif"], kurulus_yil_sayisi)
        brut_bench = dict(net_bench)
        brut_nispi = {k: round(brut_fon[k] - brut_bench[k], 1) for k in brut_fon}

        DETAY_KOLONLAR = [
            "Yillik", "3Yil_Kumulatif", "5Yil_Kumulatif", "BaslangictanBeri_Kumulatif",
            "3Yil_Yillik", "5Yil_Yillik", "BaslangictanBeri_Yillik",
        ]
        detay_ws = hedef_wb["Fon_Getiri_Analizi_Detay"]
        detay_diger_satirlar = [
            row for row in detay_ws.iter_rows(min_row=2, values_only=True)
            if row[0] is not None and row[0] != "AAV"
        ]
        detay_yeni = [
            ("AAV", tip, taraf, *[veri[k] for k in DETAY_KOLONLAR], aav_kurulus_tarihi.isoformat())
            for tip, taraf, veri in [
                ("Net", "Fon", net_fon), ("Net", "Benchmark", net_bench), ("Net", "Nispi", net_nispi),
                ("Brut", "Fon", brut_fon), ("Brut", "Benchmark", brut_bench), ("Brut", "Nispi", brut_nispi),
            ]
        ]
        detay_ws.delete_rows(2, detay_ws.max_row)
        for r, row_data in enumerate(detay_diger_satirlar + detay_yeni, start=2):
            for c, val in enumerate(row_data, start=1):
                detay_ws.cell(row=r, column=c, value=val)

        # Fon_Yillik_Getiri_Karsilastirma (Sayfa 11's TL/USD annualized 3
        # Yıl/5 Yıl bar chart) — also automated for AAV, same day. TL rows
        # are literally net_fon/net_bench's own 3Yil_Yillik/5Yil_Yillik
        # (no separate computation — same figures as the Detay table above).
        # USD rows: converts the TL cumulative return by dividing out the
        # SAME-window USD/TRY change, using a genuine raw historical USD/TRY
        # series found in Bench column 9 (found via the same row-3782
        # stray-label search used throughout this project — labeled
        # "DOLAR", history back to 1997, current value ~47.29 matching
        # Piyasa_Senaryolari's independently-sourced TCMB figure). Validated:
        # reproduces the old manual USD figures within ~0.3-1.8pp (3 Yıl:
        # computed fund 9.7%/benchmark 5.6% vs old 10.0%/7.4%; 5 Yıl:
        # computed 19.1%/14.9% vs old 19.9%/16.1%) — much closer than the
        # TL cumulative figures' own gap, since dividing by the same-basis
        # currency change cancels out most of whatever caused that gap.
        karsilastirma_yeni = []
        for yil_sayisi, donem_adi, kum_fon, kum_bench, usd_degisim in [
            (3, "Son 3 Yil", uc_yil_v, bench_uc_yil_v, usd_3yil_degisim),
            (5, "Son 5 Yil", bes_yil_v, bench_bes_yil_v, usd_5yil_degisim),
        ]:
            if kum_fon is None:
                continue
            karsilastirma_yeni.append(("AAV", donem_adi, "TL", yillandir(kum_fon, yil_sayisi), yillandir(kum_bench, yil_sayisi)))
            if usd_degisim is not None:
                usd_fon_kum = round(((1 + kum_fon / 100) / (1 + usd_degisim / 100) - 1) * 100, 2)
                usd_bench_kum = round(((1 + kum_bench / 100) / (1 + usd_degisim / 100) - 1) * 100, 2)
                karsilastirma_yeni.append(("AAV", donem_adi, "USD", yillandir(usd_fon_kum, yil_sayisi), yillandir(usd_bench_kum, yil_sayisi)))
        karsilastirma_ws = hedef_wb["Fon_Yillik_Getiri_Karsilastirma"]
        karsilastirma_diger = [
            row for row in karsilastirma_ws.iter_rows(min_row=2, values_only=True)
            if row[0] is not None and row[0] != "AAV"
        ]
        karsilastirma_ws.delete_rows(2, karsilastirma_ws.max_row)
        for r, row_data in enumerate(karsilastirma_diger + karsilastirma_yeni, start=2):
            for c, val in enumerate(row_data, start=1):
                karsilastirma_ws.cell(row=r, column=c, value=val)

    # Fon_Performans_Ek_Kriter: extra comparison columns for funds measured
    # against more than one benchmark (currently just AED — see above).
    # Fully automation-owned (unlike Fon_Performans, nothing here is ever
    # hand-entered), so every fund this dict covers gets a full replace,
    # not a merge-preserve.
    ek_kriter_tanimlari = {
        "AED": {"USD": standalone_tum.get("USD/TL"), "TL Mevduat": standalone_tum.get("KYD Mevduat Endeksi")},
        "YLC": {
            "NQXAUAGR": standalone_tum.get("Nasdaq Agriculture Ex Australia"),
            "XGIDA": standalone_tum.get("BIST Tarım Gıda Endeksi"),
        },
        "RTG": {
            "NQROBO": standalone_tum.get("Nasdaq CTA Artificial Intelligence & Robotics"),
            "XBLSM": standalone_tum.get("BIST Bilişim Getiri"),
        },
        "DGH": {"KYD ÖST Değişken": dgh_ost_degisken},
    }
    ek_ws = hedef_wb["Fon_Performans_Ek_Kriter"]
    ek_mevcut = [
        row for row in ek_ws.iter_rows(min_row=2, values_only=True)
        if row[1] is not None and row[1] not in ek_kriter_tanimlari
    ]
    ek_yeni = []
    for fon_kodu, kriterler in ek_kriter_tanimlari.items():
        for kriter_adi, pencereler in kriterler.items():
            if not pencereler:
                continue
            for pencere, deger in pencereler.items():
                ek_yeni.append((tarih_str, fon_kodu, pencere, kriter_adi, deger))
    ek_ws.delete_rows(2, ek_ws.max_row)
    for r, row_data in enumerate(ek_mevcut + ek_yeni, start=2):
        for c, val in enumerate(row_data, start=1):
            ek_ws.cell(row=r, column=c, value=val)

    # Kapak_Ozet.Tarih should track the same live date — AUM/fund counts next
    # to it stay manual (Mete's call, no automated source for those yet).
    if rapor_tarihi:
        kapak_ws = hedef_wb["Kapak_Ozet"]
        kapak_ws.cell(row=2, column=1, value=rapor_tarihi.isoformat())

    hedef_wb.save(VERI_KAYNAGI_YOLU)
    print(f"OK — rapor tarihi: {tarih_str}, {len(veriler)} fon, {len(yeni_satirlar)} otomatik satır yazıldı "
          f"({len(korunan_satirlar)} manuel satır korundu).")


if __name__ == "__main__":
    main()
