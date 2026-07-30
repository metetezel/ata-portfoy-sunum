# Peer_PE_Karsilastirma'nın 9 emsal ülkesinin (Türkiye hariç — o satıra bu
# script hiç dokunmuyor) 12 ay ileri F/K'sını artık haftalık Ata Yatırım
# e-postasından DEĞİL, doğrudan MSCI'nin ücretsiz tek-ülke factsheet'lerinden
# çekiyor. 2026-07-30'da eklendi, Mete'nin "o kısmı da otomatiğe bağlama
# şansımız var mı" sorusu üzerine.
#
# guncelle_peer_fk.py'nin YERİNE GEÇMİYOR, ONU TAMAMLIYOR — eski script hâlâ
# çalışabilir durumda (Mete PDF'i klasöre bırakmaya devam ederse), ama artık
# haftalık e-posta bağımlılığı ZORUNLU DEĞİL: bu script hem değerleri HEM
# ülke isimlerini kendi başına, sıfırdan üretiyor (eski script sadece
# değerleri güncelliyordu, isimler PDF'in çizim katmanında olduğu için hep
# elle taşınıyordu — artık isimler de bu script'in kendi ülke listesinden
# geliyor, taşımaya gerek yok).
#
# Kaynak keşfi: MSCI World/EM factsheet'i (guncelle_degerleme_fk.py) için
# zaten bilinen msci.com/documents/10199/255599/ altındaki tek-ülke
# factsheet'leri — aynı "FUNDAMENTALS" tablosu + "P/E Fwd" sütunu her
# ülkede de var, sadece dosya adı sonek kalıbı ülkeden ülkeye değişiyor
# (-price / -net / -usd-net / sonek yok) — WebSearch ile tek tek doğrulandı,
# aşağıdaki PEER_ULKELER sözlüğünde sabitlendi.
#
# ⚠️ Bilinen sınırlama (Bloomberg-kaynaklı MSCI World/EM otomasyonuyla aynı
# kategori): MSCI bu factsheet'leri ayda bir güncelliyor, rakamlar Ata
# Yatırım'ın Bloomberg tahminleriyle birebir eşleşmeyebilir (~%5-10 sapma
# beklenir, farklı kapsam/metodoloji). Kaynak notu bu yüzden sayfada
# "Bloomberg Tahminleri" yerine "MSCI" olarak değişti.
import datetime
import os
import re

import fitz
import openpyxl
import requests

VERI_KAYNAGI_YOLU = os.path.join(os.path.dirname(__file__), "Veri_Kaynagi.xlsx")
TARAYICI_UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/124.0 Safari/537.36"
)

AY_ADLARI_KISA = {
    "JAN": 1, "FEB": 2, "MAR": 3, "APR": 4, "MAY": 5, "JUN": 6,
    "JUL": 7, "AUG": 8, "SEP": 9, "OCT": 10, "NOV": 11, "DEC": 12,
}

# Türkçe görünen ad -> (MSCI factsheet TAM URL'i, MSCI'nin kendi tablo
# etiketi kelime kelime). Ata Yatırım'ın orijinal grafiğindeki AYNI 9 ülke —
# sadece kaynak değişti, seçim/kapsam Ata Yatırım'ınkiyle birebir aynı kalsın
# diye bilerek değiştirilmedi.
#
# ⚠️ URL kalıbı ülkeden ülkeye tutarlı DEĞİL — çoğu documents/10199/255599/
# altında ama Güney Afrika'nın O klasördeki dosyası donuk/eski çıktı (2013
# verisi!) — güncel (2026) olanı tamamen farklı bir yol altında
# (resources/factsheets/index_fact_sheet/) bulundu. Her ülke tam URL ile
# saklanıyor, tek bir taban+sonek varsayımı kurulmuyor.
PEER_ULKELER = {
    "Hindistan": ("https://www.msci.com/documents/10199/255599/msci-india-index-price.pdf", ["MSCI", "India"]),
    "Tayland": ("https://www.msci.com/documents/10199/255599/msci-thailand-index-price.pdf", ["MSCI", "Thailand"]),
    "Peru": ("https://www.msci.com/documents/10199/255599/msci-peru-index-net.pdf", ["MSCI", "Peru"]),
    "Meksika": ("https://www.msci.com/documents/10199/255599/msci-mexico-index-usd-net.pdf", ["MSCI", "Mexico"]),
    "Filipinler": ("https://www.msci.com/documents/10199/255599/msci-philippines-index-price.pdf", ["MSCI", "Philippines"]),
    "Kolombiya": ("https://www.msci.com/documents/10199/255599/msci-colombia-index-net.pdf", ["MSCI", "Colombia"]),
    "Endonezya": ("https://www.msci.com/documents/10199/255599/msci-indonesia-index-net.pdf", ["MSCI", "Indonesia"]),
    # ⚠️ documents/10199/255599/msci-south-africa-index-net.pdf DONUK (2013
    # verisi taşıyor) — bu URL'i tekrar denemeyin, aşağıdaki güncel.
    "Güney Afrika": ("https://www.msci.com/resources/factsheets/index_fact_sheet/msci-south-africa-net-usd.pdf", ["MSCI", "South", "Africa"]),
    "Brezilya": ("https://www.msci.com/documents/10199/255599/msci-brazil-index.pdf", ["MSCI", "Brazil"]),
}


def msci_ulke_fk_oku(url, etiket_parcalari):
    # Aynı koordinat-bazlı çıkarım tekniği guncelle_degerleme_fk.py'nin
    # msci_world_em_fk_oku()'sunda — bu PDF'lerde de tablo çizgisi yok,
    # "P/E" sütun başlığının gerçek x-konumuna en yakın sayısal değeri
    # buluyoruz (satırdaki sıraya güvenmiyoruz).
    pdf_bytes = requests.get(url, headers={"User-Agent": TARAYICI_UA}, timeout=20).content
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    page = doc[0]
    kelimeler = page.get_text("words")

    fundamentals_y = next(w[1] for w in kelimeler if w[4] == "FUNDAMENTALS")

    tarih_kelimeleri = [w[4] for w in kelimeler if abs(w[1] - fundamentals_y) <= 1.5 and w[0] > 400]
    ay_kisa = next(t.strip("(") for t in tarih_kelimeleri if t.strip("(").upper() in AY_ADLARI_KISA)
    gun = int(next(t.strip(",") for t in tarih_kelimeleri if t.rstrip(",").isdigit()))
    yil = int(next(t.strip(")") for t in tarih_kelimeleri if t.strip(")").isdigit() and len(t.strip(")")) == 4))
    veri_tarihi = datetime.date(yil, AY_ADLARI_KISA[ay_kisa.upper()], gun)

    # ⚠️ Bu sayfa "12 Aylık Tahmini F/K" (forward P/E) gösteriyor — sayfanın
    # DİĞER F/K grafiği ("BIST/GOP/Dünya Karşılaştırması", guncelle_degerleme_fk.py)
    # bilerek düz/trailing "P/E" kullanıyor, ama bu grafiğin başlığı açıkça
    # "12 Aylık Tahmini" dediği için burada "P/E Fwd" sütunu gerekiyor —
    # farklı bir metrik, aynı fonksiyonu kopyalarken min()'i max()'a çevirmeyi
    # unutup ilk denemede yanlışlıkla düz P/E okunduğu bulundu (Hindistan
    # 23,36 çıktı — gerçek P/E Fwd'i 20,01'di). "P/E" başlık satırında İKİ
    # kez geçiyor (sütunun kendisi + "P/E Fwd"nin ilk kelimesi) — sağdaki
    # (büyük x) olan "P/E Fwd"dir.
    pe_header_y = min(w[1] for w in kelimeler if w[4] == "P/E" and w[1] > fundamentals_y)
    pe_x = max(w[0] for w in kelimeler if w[4] == "P/E" and abs(w[1] - pe_header_y) <= 1.5)

    ilk = etiket_parcalari[0]
    for aday in (w for w in kelimeler if w[4] == ilk and w[1] > fundamentals_y):
        y = aday[1]
        satir = sorted((w for w in kelimeler if abs(w[1] - y) <= 1.5), key=lambda w: w[0])
        etiket_metni = [w[4] for w in satir if w[0] < 400]
        if etiket_metni[: len(etiket_parcalari)] != etiket_parcalari:
            continue
        sayisal = [w for w in satir if re.match(r"^-?\d+\.\d+$", w[4])]
        if not sayisal:
            continue
        en_yakin = min(sayisal, key=lambda w: abs(w[0] - pe_x))
        return veri_tarihi, float(en_yakin[4])

    raise RuntimeError(f"FUNDAMENTALS tablosunda '{' '.join(etiket_parcalari)}' satırı bulunamadı — PDF formatı değişmiş olabilir.")


def main():
    sonuclar = {}
    hatalar = []
    for ulke, (url, etiket) in PEER_ULKELER.items():
        try:
            tarih, fk = msci_ulke_fk_oku(url, etiket)
            sonuclar[ulke] = (tarih, fk)
            print(f"  {ulke}: {fk} ({tarih})")
        except Exception as e:
            hatalar.append(f"{ulke}: {e}")
            print(f"  {ulke}: HATA — {e}")

    if not sonuclar:
        raise RuntimeError("Hiçbir ülke okunamadı, sayfa formatı toplu değişmiş olabilir.")

    wb = openpyxl.load_workbook(VERI_KAYNAGI_YOLU)
    ws = wb["Peer_PE_Karsilastirma"]

    # Sütun düzeni (bkz. sayfanın kendi başlık satırı): A boş, B=Tarih,
    # C=Ulke_Endeks, D=FK_Orani_12Ay_Ileri — yani 0-index'te row[1]/row[2]/row[3].
    guncellenen = set()
    for row in ws.iter_rows(min_row=2):
        tarih_hucre, ulke_hucre, fk_hucre = row[1], row[2], row[3]
        ulke = ulke_hucre.value
        if ulke in sonuclar:
            tarih, fk = sonuclar[ulke]
            tarih_hucre.value = tarih.isoformat()
            fk_hucre.value = round(fk, 2)
            guncellenen.add(ulke)

    # Sheet'te henüz hiç satırı olmayan bir ülke varsa (ör. ilk çalıştırma,
    # ya da Ata Yatırım'ın grafiğinde hiç yer almamış yeni bir ülke) sona ekle.
    son_dolu_satir = ws.max_row
    for ulke, (tarih, fk) in sonuclar.items():
        if ulke in guncellenen:
            continue
        son_dolu_satir += 1
        ws.cell(row=son_dolu_satir, column=2, value=tarih.isoformat())
        ws.cell(row=son_dolu_satir, column=3, value=ulke)
        ws.cell(row=son_dolu_satir, column=4, value=round(fk, 2))
        print(f"  {ulke}: (yeni satır) -> {round(fk, 2)}")

    wb.save(VERI_KAYNAGI_YOLU)
    print(f"OK — Peer_PE_Karsilastirma güncellendi ({len(sonuclar)}/{len(PEER_ULKELER)} ülke). Türkiye satırına dokunulmadı.")
    if hatalar:
        print("UYARI — bazı ülkeler okunamadı:")
        for h in hatalar:
            print(f"  {h}")


if __name__ == "__main__":
    main()
