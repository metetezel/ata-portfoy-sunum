# Bist_Yillik_Getiriler sekmesini (Sayfa 26'nın 1989-2026 yıllık BIST getirisi
# bar grafiği) otomatik günceller — Yahoo Finance'in ücretsiz XU100.IS geçmiş
# veri serisinden (aynı kaynak build_endeks_serisi.py'nin Sayfa 25 için
# kullandığı).
#
# ⚠️ Sadece 1998 VE SONRASI yıllar güncellenir/üzerine yazılır — 1989-1997
# (Yahoo'da veri yok, XU100.IS 1997-07'den başlıyor ve 1997'nin kendi getirisi
# için 1996 yıl-sonu kapanışı gerekiyor) hâlâ elle girilmiş, DOKUNULMUYOR.
# Bu yıllar zaten sabit/değişmeyen tarihi veri, otomatikleştirmenin bir
# faydası yok.
#
# Doğrulandı (2026-07-30): 1998-2025 arası HER yıl için bu hesaplama, legacy
# PDF'ten elle girilmiş mevcut değerlerle yuvarlama farkı dışında birebir
# eşleşti (28/28 yıl) — Yahoo'nun XU100.IS'i gerçekten aynı şeyi ölçüyor.
# 2026 (içinde bulunulan, henüz kapanmamış yıl) her çalıştırmada güncel
# yılbaşından-bugüne getiriyi verir — artık senede bir elle güncellemeye
# hiç gerek yok, script her çalıştığında otomatik doğru çıkar.
import datetime
import os

import openpyxl

from build_endeks_serisi import yahoo_gunluk_seri_oku

VERI_KAYNAGI_YOLU = os.path.join(os.path.dirname(__file__), "Veri_Kaynagi.xlsx")
ILK_HESAPLANABILIR_YIL = 1998  # 1997 yıl-sonu kapanışı Yahoo'da yok, 1998'den itibaren hesaplanabilir


def yillik_getirileri_hesapla():
    seri = yahoo_gunluk_seri_oku("XU100.IS", baslangic=datetime.date(1997, 1, 1))
    tarihler = sorted(seri)
    bugun = datetime.date.today()

    def yil_sonu_kapanisi(yil):
        adaylar = [d for d in tarihler if d.year == yil]
        if not adaylar:
            return None
        return seri[max(adaylar)]

    getiriler = {}
    for yil in range(ILK_HESAPLANABILIR_YIL, bugun.year + 1):
        onceki = yil_sonu_kapanisi(yil - 1)
        simdi = yil_sonu_kapanisi(yil)
        if onceki is None or simdi is None:
            continue
        getiriler[yil] = round((simdi / onceki - 1) * 100)
    return getiriler


def main():
    getiriler = yillik_getirileri_hesapla()

    wb = openpyxl.load_workbook(VERI_KAYNAGI_YOLU)
    ws = wb["Bist_Yillik_Getiriler"]

    guncellenen = {}
    for row in ws.iter_rows(min_row=2):
        yil_hucre, deger_hucre = row[0], row[1]
        yil = yil_hucre.value
        if isinstance(yil, (int, float)) and int(yil) in getiriler:
            eski = deger_hucre.value
            yeni = getiriler[int(yil)]
            deger_hucre.value = yeni
            guncellenen[int(yil)] = (eski, yeni)

    # Sekmede henüz satırı olmayan yıllar (ör. yeni bir yıl başladığında) sona eklenir.
    son_dolu_satir = ws.max_row
    for yil, yeni in sorted(getiriler.items()):
        if yil in guncellenen:
            continue
        son_dolu_satir += 1
        ws.cell(row=son_dolu_satir, column=1, value=yil)
        ws.cell(row=son_dolu_satir, column=2, value=yeni)
        guncellenen[yil] = (None, yeni)

    wb.save(VERI_KAYNAGI_YOLU)

    print(f"OK — Bist_Yillik_Getiriler güncellendi ({ILK_HESAPLANABILIR_YIL}-{max(getiriler)}, XU100.IS/Yahoo Finance):")
    for yil in sorted(guncellenen):
        eski, yeni = guncellenen[yil]
        print(f"  {yil}: {eski} -> {yeni}")
    print("NOT: 1989-1997 arası bu script tarafından DEĞİŞTİRİLMEDİ (Yahoo'da veri yok), hâlâ manuel/sabit.")


if __name__ == "__main__":
    main()
