# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

TEAL = "00677F"
TEAL_LIGHT = "DCEEF1"
AMBER = "DE6F00"
AMBER_LIGHT = "FBEADA"
GREY_LIGHT = "F2F2F2"

header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor=TEAL)
title_font = Font(name="Calibri", bold=True, color=TEAL, size=14)
note_font = Font(name="Calibri", italic=True, color="666666", size=10)
thin = Side(style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = openpyxl.Workbook()
wb.remove(wb.active)


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[row].height = 28


def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_note(ws, cell_ref, text, fill=None):
    ws[cell_ref] = text
    ws[cell_ref].font = note_font
    if fill:
        ws[cell_ref].fill = PatternFill("solid", fgColor=fill)


# ---------------------------------------------------------------
# 0. TALIMATLAR
# ---------------------------------------------------------------
ws = wb.create_sheet("Talimatlar")
ws.sheet_view.showGridLines = False
ws["B2"] = "Veri_Kaynagi.xlsx — Ata Portföy Haftalık Sunum Veri Kaynağı"
ws["B2"].font = Font(name="Calibri", bold=True, color=TEAL, size=16)
ws["B4"] = (
    "Bu dosya, haftalık Sunum.xlsx'in yerini alacak otomatik sistemin okuyacağı TEK veri kaynağıdır. "
    "Amaç: dağınık 33 sekmeli/663 sütunlu yapı yerine, her hafta sadece bu dosyadaki ilgili "
    "hücrelere veri girmeniz/güncellemeniz yeterli olsun."
)
ws["B4"].alignment = Alignment(wrap_text=True)
ws.merge_cells("B4:H4")
ws.row_dimensions[4].height = 45

rows = [
    ("Sekme", "İçerik", "Kaynak", "Güncelleme Sıklığı"),
    ("Fon_Katalog", "Fon kodu, adı, türü, risk, ücret, vergi (referans tablo)", "Manuel — nadiren", "Fon eklendiğinde/çıkarıldığında"),
    ("Fon_Performans", "Fon getirileri (aylık/3A/6A/yıllık/YBB/3Y/5Y)", "Rasyonet (API/eklenti)", "Haftalık"),
    ("Fon_Performans_Ek_Kriter", "Birden fazla karşılaştırma kriteri olan fonlar için ek getiri kolonları (ör. AED: USD, TL Mevduat)", "Rasyonet (API/eklenti)", "Haftalık"),
    ("Fon_Fiyat_Serisi", "Fon ve benchmark fiyat serisi (çizgi grafik verisi)", "Rasyonet", "Haftalık"),
    ("Fon_Portfoy_Dagilimi", "Fon varlık dağılımı % (pasta grafik verisi)", "TEFAS API (otomatik, guncelle_fon_portfoy_dagilimi.py)", "Günlük"),
    ("Fon_Getiri_Analizi_Detay", "Genişletilmiş getiri analizi (net/brüt, sadece bazı fonlarda)", "Rasyonet", "Haftalık"),
    ("Fon_Yillik_Getiri_Karsilastirma", "3/5 yıllık TL-USD yıllıklandırılmış getiri (sadece bazı fonlarda)", "Rasyonet", "Haftalık"),
    ("Fon_Temettu_Verimi", "Dönemlere göre temettü verimi % (sadece temettü ödeyen fonlarda, ör. AYA)", "Manuel — çeyreklik dağıtım sonrası", "Çeyreklik (Ocak/Nisan/Temmuz/Ekim)"),
    ("Piyasa_Senaryolari", "BIST/USD/EUR/S&P500/EuroStoxx senaryo tahminleri", "Bloomberg (elle not)", "Haftalık"),
    ("Makro_Gostergeler", "GSYH, enflasyon, kur, dış ticaret vb. (2021-2026T)", "Bloomberg konsensus", "Aylık/Çeyreklik"),
    ("Degerleme_CDS_MSCI", "BIST F/K, Türkiye CDS, MXWO/MXEF endeksleri", "Bloomberg (remote terminal akışı)", "Haftalık"),
    ("Bist_Yillik_Getiriler", "BIST-100'ün 1989'dan beri yıllık TL getirisi", "Legacy PDF'ten elle girildi", "Yılda 1 kez (yıl kapanınca)"),
    ("Peer_PE_Karsilastirma", "Pazartesi gelen araştırma mailindeki 10 ülke/endeks F/K değeri", "Ata Yatırım araştırma maili (elle, 10 sayı)", "Haftalık (Pazartesi)"),
    ("Fon_Metinleri", "Her fonun sabit açıklama metni", "Manuel — nadiren", "Fon stratejisi değişince"),
    ("Sabit_Metinler", "Kurumsal tanıtım, 100 TL stratejisi, vergi metinleri", "Manuel — nadiren", "İçerik değişince"),
    ("Fon_Avantajlari_Maddeler", "Sayfa 28'in iki madde listesi", "Manuel — nadiren", "İçerik değişince"),
    ("Vergi_Bireysel_Fon_Listesi", "Sayfa 29'un %15 stopaja tabi olmayan fonlar listesi", "Manuel — nadiren", "Mevzuat değişince"),
    ("Vergi_ANZ_Tablosu", "Sayfa 29'un ANZ vergisel avantajlar tablosu", "Manuel — nadiren", "Mevzuat değişince"),
    ("Ekip", "İletişim sayfası kişi kartları", "Manuel — nadiren", "Personel değişince"),
    ("Ekip_Organizasyon", "Fon Yönetimi / Fon Operasyon iç organizasyon şeması", "Manuel — nadiren", "Personel/departman değişince"),
    ("Ata_Fonlari_Performans_Ozet", "\"Ata Fonları: Güçlü Performans\" sayfasının 1 yıllık getiri karşılaştırması", "Rasyonet (otomatik)", "Haftalık"),
    ("Ata_Fonlari_Performans_Ozet_YBB", "Aynı sayfanın yılbaşından beri (YBB) versiyonu — farklı fon/gösterge seçimi", "Rasyonet (otomatik)", "Haftalık"),
    ("Yatirim_100TL_Dagilim", "\"100 TL Yatırımınızı Nasıl Değerlendirmenizi Öneriyoruz?\" pasta grafiği", "Manuel — nadiren", "Strateji değişince"),
    ("Yatirim_100TL_Maddeler", "Aynı sayfanın 4 alt-madde açıklaması + TL tutarı", "Manuel — nadiren", "Strateji değişince"),
    ("Kapak_Ozet", "Toplam AUM ve fon sayısı", "Manuel", "Haftalık"),
]
start_row = 7
for r, row_data in enumerate(rows):
    for c, val in enumerate(row_data):
        cell = ws.cell(row=start_row + r, column=2 + c, value=val)
        cell.border = border
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        if r == 0:
            cell.font = header_font
            cell.fill = header_fill
        else:
            cell.font = Font(name="Calibri", size=10.5)
autosize(ws, [3, 22, 46, 26, 24])

legend_row = start_row + len(rows) + 3
ws.cell(row=legend_row, column=2, value="Renk kodu:").font = Font(bold=True, size=10.5)
add_note(ws, f"C{legend_row}", "  Rasyonet'ten otomatik/yarı-otomatik beslenecek alanlar  ", TEAL_LIGHT)
add_note(ws, f"C{legend_row + 1}", "  Bloomberg / manuel giriş gereken alanlar  ", AMBER_LIGHT)

ws.sheet_properties.tabColor = TEAL

# ---------------------------------------------------------------
# 1. FON_KATALOG
# ---------------------------------------------------------------
ws = wb.create_sheet("Fon_Katalog")
headers = ["Fon_Kodu", "Fon_Adi", "Kisa_Ad", "Fon_Turu", "Risk_Degeri_1_7", "Yonetim_Ucreti", "Stopaj", "Benchmark_Aciklama", "Benchmark_Kisa_Ad", "Benchmark2_Kisa_Ad", "Not"]
for c, h in enumerate(headers, start=1):
    ws.cell(row=1, column=c, value=h)
style_header(ws, 1, len(headers))

# Kisa_Ad = short possessive-ready name used inside narrative headline callouts
# (e.g. "İkinci Hisse Fon'umuz, 2012'den beri..."), separate from the full
# legal Fon_Adi used everywhere else.
#
# Benchmark_Kisa_Ad = short display name for table headers, filled in
# manually rather than regex-derived from Benchmark_Aciklama — a first
# attempt tried extracting this with a regex assuming the benchmark name
# was a leading all-caps run (worked for AAV's "BIST 100 Getiri Endeksi",
# silently failed for AYA's "BIST Temettü 25 Getiri Endeksi" since "Temettü"
# breaks the all-caps assumption, falling back to a generic "Benchmark"
# label). Composite/blended benchmarks (URA/JET/YLC/RTG) can't be
# meaningfully auto-shortened at all — left blank until each fund's page
# is actually built, same incremental-fill convention as Fon_Metinleri etc.
#
# Benchmark2_Kisa_Ad = a SECOND comparison line's chart legend name, only
# used by a fund with two benchmarks at once (currently just AED). Kept
# separate from Fon_Performans_Ek_Kriter's own Kriter_Adi column since the
# legacy deck uses a DIFFERENT short name for the same series in the table
# ("TL Mevduat") vs. the chart legend ("KYD Mevduat Endeksi") — not
# reusable from one place to the other.
fon_katalog_data = [
    ("AYA", "ATA Kâr Payı Ödeyen Hisse Senedi (TL) Fonu", "Kâr Payı Ödeyen Hisse Fon", "Hisse Senedi Fonu", 6, "%2,50", "%0", "BIST Temettü 25 Getiri Endeksi", "BIST Temettü 25", "", ""),
    ("AAV", "ATA İkinci Hisse Senedi (TL) Fonu", "İkinci Hisse Fon", "Hisse Senedi Fonu", 6, "%2,25", "%0", "BIST 100 Getiri Endeksi", "BIST 100", "", ""),
    ("TLZ", "ATA Katılım Hisse Senedi (TL) Fonu", "Katılım Hisse Fon", "Hisse Senedi Fonu", 6, "%2,25", "%0", "BIST Katılım 100 Getiri (XK100)", "BIST Katılım 100", "", ""),
    ("URA", "ATA Enerji Değişken Fon", "Enerji Değişken Fon", "Değişken Fon", 6, "%2,50", "%17,5", "%65 Solactive Global Uranium & Nuclear TR Index + %35 BIST Elektrik Getiri Endeksi", "Benchmark", "", ""),
    ("JET", "ATA Havacılık ve Savunma Teknolojileri Değişken Fonu", "Havacılık ve Savunma Fon", "Değişken Fon", 6, "%2,50", "%17,5", "%65 Nasdaq US Aerospace & Defense TR Index + %35 BIST Teknoloji Ağ. Sın. Getiri End.", "Benchmark", "", ""),
    ("YLC", "ATA Tarım ve Gıda Değişken Fon", "Tarım ve Gıda Fon", "Değişken Fon", 6, "%1,75", "%17,5", "%65 NASDAQ Global Ex-Australia Agriculture (NQXAUAGR) + %35 XGIDA", "Benchmark", "", ""),
    ("RTG", "ATA Robotik Teknolojileri Değişken Fon", "Robotik Teknolojiler Fon", "Değişken Fon", 6, "%2,50", "%17,5", "%65 NASDAQ CTA AI & Robotics (NQROBO) + %35 XBLSM", "Benchmark", "", ""),
    ("AED", "ATA Birinci Değişken Fon", "Birinci Değişken Fon", "Değişken Fon", 5, "%1,75", "%17,5*", "BIST-100 Getiri Endeksi / KYD Mevduat Endeksi", "BIST-100 Getiri", "KYD Mevduat Endeksi", "*1 yıl elde tutulursa %0"),
    ("AAS", "ATA Fon Sepeti Serbest Fon", "Fon Sepeti Serbest Fon", "Serbest Fon", 5, "%0,75", "%17,5", "S&P 500 (ETF bazlı)", "", "", ""),  # benchmarksız — legacy tek "Fon" kolonu gösteriyor
    ("ANZ", "ATA Dördüncü Serbest (Döviz) Fon", "Dördüncü Serbest Fon", "Serbest Fon", 6, "%0,75", "%17,5", "KYD USD Mevduat", "KYD USD Mevduat", "", "Eurobond fonu"),
    ("DGH", "ATA Para Piyasası Serbest (TL) Fon", "Para Piyasası Serbest Fon", "Serbest Fon", 1, "%0,95", "%17,5", "KYD Mevduat Endeksi", "KYD Mevduat", "", ""),
    ("AAL", "ATA Portföy Para Piyasası (TL) Fonu", "Para Piyasası Fon", "Para Piyasası Fonu", 1, "%1,00", "%17,5", "", "", "", ""),
    ("PKF", "ATA Altın Katılım Fonu", "Altın Katılım Fon", "Katılım Fonu", 6, "%1,00", "%17,5", "", "Benchmark", "", ""),
    ("PKP", "ATA Katılım Para Piyasası Serbest (TL) Fon", "Katılım Para Piyasası Fon", "Serbest Fon", 1, "%0,80", "%17,5", "", "", "", ""),
]
for r, row_data in enumerate(fon_katalog_data, start=2):
    for c, val in enumerate(row_data, start=1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.border = border
        cell.alignment = Alignment(vertical="center", wrap_text=(c in (8, 11)))
autosize(ws, [10, 42, 26, 18, 14, 12, 9, 55, 18, 20, 22])
ws.freeze_panes = "A2"
ws.sheet_properties.tabColor = GREY_LIGHT

FON_KODU_RANGE = "Fon_Katalog!$A$2:$A$15"

# ---------------------------------------------------------------
# helper to build a tidy weekly sheet with fon_kodu dropdown
# ---------------------------------------------------------------

def build_tidy_sheet(name, headers, widths, example_rows, tab_color=TEAL_LIGHT, fonkodu_col=None, dropdown_cols=None, ornek_notu=True):
    ws = wb.create_sheet(name)
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, len(headers))
    for r, row_data in enumerate(example_rows, start=2):
        for c, val in enumerate(row_data, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = border
    if example_rows and ornek_notu:
        note_row = 2 + len(example_rows)
        ws.cell(row=note_row, column=1, value="↑ örnek satır — silip kendi verinizi girin").font = note_font
    autosize(ws, widths)
    ws.freeze_panes = "A2"
    ws.sheet_properties.tabColor = tab_color
    if fonkodu_col:
        dv = DataValidation(type="list", formula1=f"={FON_KODU_RANGE}", allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{get_column_letter(fonkodu_col)}2:{get_column_letter(fonkodu_col)}2000")
    if dropdown_cols:
        for col, options in dropdown_cols.items():
            dv = DataValidation(type="list", formula1='"' + ",".join(options) + '"', allow_blank=True)
            ws.add_data_validation(dv)
            dv.add(f"{get_column_letter(col)}2:{get_column_letter(col)}2000")
    return ws


# ---------------------------------------------------------------
# 2. FON_PERFORMANS
# ---------------------------------------------------------------
PENCERE_OPTS = ["Aylik", "3 Aylik", "6 Aylik", "Yillik", "Yilbasindan Beri", "3 Yil", "5 Yil", "Kurulustan Beri"]
build_tidy_sheet(
    "Fon_Performans",
    ["Tarih", "Fon_Kodu", "Pencere", "Fon_Getiri_Yuzde", "Benchmark_Getiri_Yuzde"],
    [12, 10, 16, 16, 18],
    [
        ("2026-07-27", "AAV", "Aylik", -2.6, -2.7),
        ("2026-07-27", "AAV", "3 Aylik", -4.9, -3.5),
        ("2026-07-27", "AAV", "Yillik", 18.3, 33.6),
        ("2026-07-27", "AAV", "Yilbasindan Beri", 15.8, 25.2),
        ("2026-07-27", "AAV", "3 Yil", 157.9, 133.4),
        ("2026-07-27", "AAV", "5 Yil", 1298.3, 1072.2),
        ("2026-07-27", "AAV", "Kurulustan Beri", 5924.4, 2694.2),
        # AYA: computed from the validated price series (data/aya_price_series.json —
        # Price!H fund / Bench!38 benchmark, see build_fiyat_serisi.py), NOT
        # transcribed from the legacy PDF — that table's own "Kuruluştan Beri" mixes
        # a NET fund figure with what looks like a raw benchmark index value read
        # directly as if it were the return %, a real inconsistency found this
        # session. Fund is since its own 2010-06-17 start; benchmark is since ITS
        # own start (2011-07-01, when BIST Temettü 25 began being calculated) —
        # not a true "since fund inception" figure for the benchmark side, since
        # the index simply doesn't exist before then.
        ("2026-07-29", "AYA", "Kurulustan Beri", 4554.8, 4994.2),
    ],
    fonkodu_col=2,
    dropdown_cols={3: PENCERE_OPTS},
)

# ---------------------------------------------------------------
# 2b. FON_PERFORMANS_EK_KRITER — for funds compared against MORE than one
#     benchmark at once (currently just AED: BIST-100 Getiri / USD / TL
#     Mevduat, vs. the single Benchmark_Getiri_Yuzde column every other
#     fund uses). Fully automation-owned, like Fon_Performans's short-term
#     windows — guncelle_fon_getirileri.py replaces every row for a given
#     fund on each run, no manual-preserve logic needed here (unlike
#     Fon_Performans's 3 Yıl/5 Yıl/Kuruluştan Beri exception).
# ---------------------------------------------------------------
build_tidy_sheet(
    "Fon_Performans_Ek_Kriter",
    ["Tarih", "Fon_Kodu", "Pencere", "Kriter_Adi", "Getiri_Yuzde"],
    [12, 10, 16, 14, 16],
    [
        ("2026-07-27", "AED", "Aylik", "USD", 1.6),
        ("2026-07-27", "AED", "Aylik", "TL Mevduat", 9.7),
        ("2026-07-27", "YLC", "Aylik", "NQXAUAGR", 8.2),
        ("2026-07-27", "YLC", "Aylik", "XGIDA", 7.3),
        ("2026-07-27", "RTG", "Aylik", "NQROBO", -0.6),
        ("2026-07-27", "RTG", "Aylik", "XBLSM", 12.0),
        ("2026-07-27", "DGH", "Aylik", "KYD ÖST Değişken", 3.4),
    ],
    fonkodu_col=2,
    dropdown_cols={3: PENCERE_OPTS},
)

# ---------------------------------------------------------------
# 3. FON_FIYAT_SERISI
# ---------------------------------------------------------------
build_tidy_sheet(
    "Fon_Fiyat_Serisi",
    ["Tarih", "Fon_Kodu", "Fon_Fiyat_Endeks", "Benchmark_Fiyat_Endeks"],
    [12, 10, 20, 22],
    [("2026-07-20", "AAV", 7180, 2650), ("2026-07-27", "AAV", 7553.3, 2694.2)],
    fonkodu_col=2,
)

# ---------------------------------------------------------------
# 4. FON_PORTFOY_DAGILIMI
# ---------------------------------------------------------------
build_tidy_sheet(
    "Fon_Portfoy_Dagilimi",
    ["Tarih", "Fon_Kodu", "Varlik_Sinifi", "Yuzde"],
    [12, 10, 30, 10],
    [
        ("2026-07-27", "AAV", "Yerli Hisse Senedi", 98),
        ("2026-07-27", "AAV", "Nakit Teminat", 2),
        ("2026-07-27", "AYA", "Yerli Hisse Senedi", 95),
        ("2026-07-27", "AYA", "Nakit Teminat", 5),
        ("2026-07-27", "TLZ", "Yerli Hisse Senedi", 90),
        ("2026-07-27", "TLZ", "Nakit Teminat", 10),
        ("2026-07-27", "ANZ", "Özel Sektör Dış Borçlanma Araçları", 87),
        ("2026-07-27", "ANZ", "Kamu Dış Borçlanma Araçları", 12),
        ("2026-07-27", "ANZ", "Nakit Teminat", 1),
        ("2026-07-27", "PKP", "BİST Taahütlü İşlem Pazarı Alımı", 65),
        ("2026-07-27", "PKP", "Özel Sektör Kira Sertifikaları", 25),
        ("2026-07-27", "PKP", "Katılma Hesabı (TL)", 8),
        ("2026-07-27", "PKP", "Kamu Kira Sertifikaları (TL)", 1),
        ("2026-07-27", "AED", "Yerli Hisse Senedi", 55),
        ("2026-07-27", "AED", "Finansman Bonosu", 12),
        ("2026-07-27", "AED", "Yatırım Fonları Katılma Payları", 18),
        ("2026-07-27", "AED", "Nakit Teminat", 8),
        ("2026-07-27", "AED", "Özel Sektör Tahvili", 7),
        ("2026-07-27", "AAS", "S&P 500", 79),
        ("2026-07-27", "AAS", "TL Sabit Getirili", 21),
        ("2026-07-27", "YLC", "Yabancı Hisse Senedi", 65),
        ("2026-07-27", "YLC", "Yerli Hisse Senedi", 35),
        ("2026-07-27", "RTG", "Yabancı Hisse Senedi", 65),
        ("2026-07-27", "RTG", "Yerli Hisse Senedi", 35),
        ("2026-07-27", "JET", "Yabancı Hisse Senedi", 65),
        ("2026-07-27", "JET", "Yerli Hisse Senedi", 35),
        ("2026-07-27", "URA", "Yabancı Hisse Senedi", 80),
        ("2026-07-27", "URA", "Yerli Hisse Senedi", 20),
        ("2026-07-27", "DGH", "Mevduat (TL)", 60),
        ("2026-07-27", "DGH", "Finansman Bonosu", 25),
        ("2026-07-27", "DGH", "Ters Repo", 15),
        ("2026-07-27", "PKF", "Fiziki Altın ve Altına Dayalı Kira Sertifikaları", 80),
        ("2026-07-27", "PKF", "Kira Sertifikaları ve Faiz Geliri İçermeyen Yabancı Yatırım Fonları ve Borsa Yatırım Fonları", 20),
    ],
    fonkodu_col=2,
    ornek_notu=False,
)

# ---------------------------------------------------------------
# 4b. FON_TEMETTU_VERIMI — only dividend-paying funds (e.g. AYA) get this;
# a rarely-changing per-fund extension, exactly like Fon_Getiri_Analizi_Detay/
# Fon_Yillik_Getiri_Karsilastirma above are only populated for some funds.
# ---------------------------------------------------------------
build_tidy_sheet(
    "Fon_Temettu_Verimi",
    ["Fon_Kodu", "Sira", "Donem_Etiket", "Temettu_Verimi_Yuzde"],
    [10, 8, 16, 20],
    [
        ("AYA", 1, "2023 Temmuz", 8.33),
        ("AYA", 2, "2024 Ocak", 1.16),
        ("AYA", 3, "2024 Nisan", 1.09),
        ("AYA", 4, "2024 Temmuz", 3.32),
        ("AYA", 5, "2024 Ekim", 1.94),
        ("AYA", 6, "2025 Ocak", 0.99),
        ("AYA", 7, "2025 Nisan", 2.52),
        ("AYA", 8, "2025 Temmuz", 5.03),
        ("AYA", 9, "2025 Ekim", 0.76),
        ("AYA", 10, "2026 Ocak", 0.59),
        ("AYA", 11, "2026 Nisan", 1.83),
        ("AYA", 12, "2026 Temmuz", 1.47),
    ],
    fonkodu_col=1,
    ornek_notu=False,
)

# ---------------------------------------------------------------
# 5. PIYASA_SENARYOLARI
# ---------------------------------------------------------------
ENSTRUMAN_OPTS = ["BIST100 Endeksi", "USD/TL", "EURO/TL", "S&P 500 Endeksi", "Euro Stoxx 50 Endeksi"]
build_tidy_sheet(
    "Piyasa_Senaryolari",
    ["Tarih", "Enstruman", "Mevcut_Durum", "Kotumser", "Baz", "Iyimser"],
    [12, 20, 14, 12, 12, 12],
    [
        ("2026-07-27", "BIST100 Endeksi", 13944, 13000, 17000, 19000),
        ("2026-07-27", "USD/TL", 47.16, 54.70, 51.53, 49.90),
        ("2026-07-27", "EURO/TL", 53.70, 63.40, 59.77, 57.92),
        ("2026-07-27", "S&P 500 Endeksi", 7411, 7000, 7800, 8250),
        ("2026-07-27", "Euro Stoxx 50 Endeksi", 6361, 5800, 6300, 6800),
    ],
    dropdown_cols={2: ENSTRUMAN_OPTS},
    tab_color=AMBER_LIGHT,
    ornek_notu=False,
)

# ---------------------------------------------------------------
# 6. MAKRO_GOSTERGELER
# ---------------------------------------------------------------
MAKRO_YILLAR = [2021, 2022, 2023, 2024, 2025, 2026]
MAKRO_VERI = [
    ("GSYH (milyar TL)", [7249, 15007, 26276, 39806, 62905, 82330]),
    ("GSYH (milyar $)", [803, 906, 1108, 1214, 1595, 1741]),
    ("GSYH Buyumesi", [11.0, 5.6, 4.5, 3.2, 3.8, 4.0]),
    ("Euro/ABD $ (yil sonu)", [1.13, 1.07, 1.11, 1.04, 1.18, 1.16]),
    ("Euro/ABD $ (ortalama)", [1.18, 1.05, 1.08, 1.08, 1.13, 1.17]),
    ("ABD $/TL (yil sonu)", [12.98, 18.70, 29.44, 35.22, 42.86, 51.53]),
    ("ABD $/TL (ortalama)", [8.87, 16.57, 23.71, 32.79, 39.44, 47.28]),
    ("Euro/TL (yil sonu)", [14.68, 19.93, 32.57, 36.74, 50.45, 59.77]),
    ("Euro/TL (ortalama)", [10.45, 17.39, 25.64, 35.48, 44.65, 55.27]),
    ("UFE (yil sonu) (%)", [79.89, 97.72, 44.22, 28.52, 27.67, 20.71]),
    ("TUFE (yil sonu) (%)", [36.08, 64.27, 64.77, 44.38, 30.89, 27.50]),
    ("Cari Islemler Dengesi (milyar ABD $)", [-15, -49, -44, -10, -24, -27]),
    ("Cari Islemler Dengesi / GSYH (%)", [-1.9, -5.4, -4.0, -0.8, -1.5, -1.8]),
    ("Ihracat (fob, milyar ABD $)", [225, 254, 255, 262, 273, 285]),
    ("Ithalat (cif, milyar ABD $)", [271, 364, 362, 344, 366, 394]),
    ("Dis Ticaret Dengesi (milyar ABD $)", [-46, -110, -107, -82, -92, -109]),
    ("Brent Petrol (ABD$/varil)", [71, 99, 82, 80, 69, 60]),
]
MAKRO_SATIRLAR = [
    (yil, gosterge, deger)
    for gosterge, degerler in MAKRO_VERI
    for yil, deger in zip(MAKRO_YILLAR, degerler)
]
build_tidy_sheet(
    "Makro_Gostergeler",
    ["Yil", "Gosterge", "Deger"],
    [10, 32, 14],
    MAKRO_SATIRLAR,
    tab_color=AMBER_LIGHT,
    ornek_notu=False,
)

# ---------------------------------------------------------------
# 7. DEGERLEME_CDS_MSCI
# ---------------------------------------------------------------
METRIK_OPTS = ["Turkiye F/K", "GOP F/K", "Dunya F/K", "Turkiye CDS", "MXWO", "MXEF", "BIST100 Piyasa Degeri/GSYIH"]
build_tidy_sheet(
    "Degerleme_CDS_MSCI",
    ["Tarih", "Metrik", "Deger"],
    [12, 28, 12],
    [
        # "Turkiye F/K" (eskiden "BIST-100 F/K", CEIC kaynaklı) 2026-08-17'de
        # MSCI'nin kendi Türkiye endeksine geçti — bkz. guncelle_degerleme_fk.py.
        ("2026-07-27", "Turkiye F/K", 17.0),
        ("2026-07-27", "GOP F/K", 17.2),
        ("2026-07-27", "Dunya F/K", 22.9),
        ("2026-07-27", "Turkiye CDS", 258),
    ],
    dropdown_cols={2: METRIK_OPTS},
    tab_color=AMBER_LIGHT,
)

# ---------------------------------------------------------------
# 7a. BIST_YILLIK_GETIRILER — Sayfa 26's 38-year annual TL return bar chart.
# No live source found anywhere in the project's workbooks (checked
# COPY Fon Broşür.xlsx and Sunum.xlsx) — real, permanent historical record,
# transcribed once from the legacy PDF's own rendered chart (cross-checked
# against its raw PDF text run, which independently confirms every value).
# Only the current (in-progress) year's bar should ever need updating.
# ---------------------------------------------------------------
BIST_YILLIK_GETIRILER = [
    1989, 493, 1990, 47, 1991, 34, 1992, -8, 1993, 417, 1994, 32, 1995, 47, 1996, 144,
    1997, 254, 1998, -25, 1999, 485, 2000, -38, 2001, 46, 2002, -25, 2003, 80, 2004, 34,
    2005, 59, 2006, -2, 2007, 42, 2008, -52, 2009, 97, 2010, 25, 2011, -22, 2012, 53,
    2013, -13, 2014, 26, 2015, -16, 2016, 9, 2017, 48, 2018, -21, 2019, 25, 2020, 29,
    2021, 26, 2022, 197, 2023, 36, 2024, 32, 2025, 15, 2026, 24,
]
bist_yillik_satirlar = [
    (BIST_YILLIK_GETIRILER[i], BIST_YILLIK_GETIRILER[i + 1])
    for i in range(0, len(BIST_YILLIK_GETIRILER), 2)
]
build_tidy_sheet(
    "Bist_Yillik_Getiriler",
    ["Yil", "Getiri_Yuzde"],
    [10, 14],
    bist_yillik_satirlar,
    tab_color=AMBER_LIGHT,
    ornek_notu=False,
)

# ---------------------------------------------------------------
# 7b. FON_GETIRI_ANALIZI_DETAY — extended fund page variant (not every
#     fund has this; only render the section on the web page when rows exist)
# ---------------------------------------------------------------
ws = wb.create_sheet("Fon_Getiri_Analizi_Detay")
headers = ["Fon_Kodu", "Tip", "Taraf", "Yillik", "3Yil_Kumulatif", "5Yil_Kumulatif",
           "BaslangictanBeri_Kumulatif", "3Yil_Yillik", "5Yil_Yillik", "BaslangictanBeri_Yillik", "Kurulus_Tarihi"]
for c, h in enumerate(headers, start=1):
    ws.cell(row=1, column=c, value=h)
style_header(ws, 1, len(headers))
detay_data = [
    ("AAV", "Net", "Fon", 18.3, 157.9, 1298.3, 5924.4, 37.1, 69.5, 35.1, "2012-11-16"),
    ("AAV", "Net", "Benchmark", 33.6, 133.4, 1072.2, 2694.2, 32.7, 63.6, 27.7, "2012-11-16"),
    ("AAV", "Net", "Nispi", -15.3, 24.5, 226.0, 3230.2, 4.5, 5.9, 7.4, "2012-11-16"),
    ("AAV", "Brut", "Fon", 20.9, 172.5, 1433.6, 7553.3, 39.7, 72.6, 37.5, "2012-11-16"),
    ("AAV", "Brut", "Benchmark", 33.6, 133.4, 1072.2, 2694.2, 32.7, 63.6, 27.7, "2012-11-16"),
    ("AAV", "Brut", "Nispi", -12.7, 39.0, 361.0, 4859.1, 7.0, 9.0, 9.8, "2012-11-16"),
]
for r, row_data in enumerate(detay_data, start=2):
    for c, val in enumerate(row_data, start=1):
        ws.cell(row=r, column=c, value=val).border = border
autosize(ws, [10, 8, 12, 10, 14, 14, 24, 12, 12, 22, 14])
ws.freeze_panes = "A2"
ws.sheet_properties.tabColor = TEAL_LIGHT

# ---------------------------------------------------------------
# 7c. FON_YILLIK_GETIRI_KARSILASTIRMA — the 3yr/5yr TL-vs-USD bar charts
# ---------------------------------------------------------------
ws = wb.create_sheet("Fon_Yillik_Getiri_Karsilastirma")
headers = ["Fon_Kodu", "Donem", "Para_Birimi", "Fon_Getiri_Yillik", "Benchmark_Getiri_Yillik"]
for c, h in enumerate(headers, start=1):
    ws.cell(row=1, column=c, value=h)
style_header(ws, 1, len(headers))
yillik_karsilastirma_data = [
    ("AAV", "Son 5 Yil", "TL", 68.7, 63.3),
    ("AAV", "Son 5 Yil", "USD", 19.9, 16.1),
    ("AAV", "Son 3 Yil", "TL", 32.6, 29.4),
    ("AAV", "Son 3 Yil", "USD", 10.0, 7.4),
]
for r, row_data in enumerate(yillik_karsilastirma_data, start=2):
    for c, val in enumerate(row_data, start=1):
        ws.cell(row=r, column=c, value=val).border = border
autosize(ws, [10, 12, 12, 18, 22])
dv = DataValidation(type="list", formula1='"Son 3 Yil,Son 5 Yil"', allow_blank=True)
ws.add_data_validation(dv)
dv.add("B2:B2000")
dv2 = DataValidation(type="list", formula1='"TL,USD"', allow_blank=True)
ws.add_data_validation(dv2)
dv2.add("C2:C2000")
ws.freeze_panes = "A2"
ws.sheet_properties.tabColor = TEAL_LIGHT

# ---------------------------------------------------------------
# 8. PEER_PE_KARSILASTIRMA
# ---------------------------------------------------------------
ws = wb.create_sheet("Peer_PE_Karsilastirma")
ws.sheet_view.showGridLines = False
ws["B2"] = "Haftalık Peer F/K Karşılaştırması"
ws["B2"].font = title_font
ws["B4"] = (
    "Kaynak: her Pazartesi gelen \"Ata Yatırım Araştırma — Türkiye ve Diğer Ülkeler Karşılaştırma\" maili, "
    "2. sayfa, \"12 Ay İleri F/K Oranı - Türkiye ve Benzer Ülkeler\" grafiği. Bu grafiğin sayıları PDF'te "
    "metin olarak yer alıyor ama ülke/endeks isimleri grafiğin çizimi içinde olduğu için otomatik "
    "eşleştirilemiyor — aşağıya her hafta 10 çubuğu solaçtan sağa, isim ve değer olarak elle girin "
    "(görsel yapıştırmaya göre daha sağlam ve web tarafında native grafiğe dönüşür)."
)
ws["B4"].alignment = Alignment(wrap_text=True)
ws.merge_cells("B4:H4")
ws.row_dimensions[4].height = 70
headers = ["Tarih", "Ulke_Endeks", "FK_Orani_12Ay_Ileri"]
for c, h in enumerate(headers, start=2):
    cell = ws.cell(row=7, column=c, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = border
# Real 27 Temmuz 2026 values, read directly off the rendered legacy PDF
# page 24 chart (bars left-to-right) — F/K values are real extractable PDF
# text, but the country/index NAMES are drawn inside the chart graphic
# itself and aren't text-extractable, so they're transcribed by eye here
# (matches the sheet's own weekly hand-entry instructions in B4 above).
example_vals = [19.8, 18.1, 13.7, 12.9, 10.6, 9.8, 9.3, 8.6, 8.4, 0.2]
example_isimler = [
    "Hindistan", "Tayland", "Peru", "Meksika", "Filipinler",
    "Kolombiya", "Endonezya", "Güney Afrika", "Brezilya", "Türkiye",
]
for i, (v, isim) in enumerate(zip(example_vals, example_isimler)):
    r = 8 + i
    ws.cell(row=r, column=2, value="2026-07-27" if i == 0 else None).border = border
    ws.cell(row=r, column=3, value=isim).border = border
    ws.cell(row=r, column=4, value=v).border = border
ws.cell(row=8 + len(example_vals), column=3, value="↑ her hafta yeni maildeki sırayla üzerine yazılmalı").font = note_font
autosize(ws, [3, 14, 26, 20])
ws.freeze_panes = "A8"
ws.sheet_properties.tabColor = AMBER

# ---------------------------------------------------------------
# 9. FON_METINLERI
# ---------------------------------------------------------------
ws = wb.create_sheet("Fon_Metinleri")
ws["A1"] = "Fon_Kodu"
ws["B1"] = "Aciklama_Metni"
style_header(ws, 1, 2)

fon_metinleri = [
    ("AYA", "Fon, 11 Nisan 2022 tarihinde yeni stratejisiyle işlem görmeye başlamıştır. Ocak, Nisan, Temmuz ve Ekim aylarının başında olmak üzere yılda dört kez yatırımcılara nakit temettü ödemektedir. İçinde yer alan hisseler piyasa koşullarına ve temettü verimliliğine göre güncellenmektedir. Fon'un karşılaştırma ölçütü BIST-Temettü 25 Getiri Endeksi'dir. Stopaj oranı %0'dır."),
    ("AAV", "Fon, ağırlıklı BIST100'de işlem gören ve kar büyüme potansiyeli olan 30-35 şirkete yatırım yapar. Fon yönetiminde aktif yönetim modeli uygulanır. Fon'un karşılaştırma ölçütü BIST-100 Getiri Endeksi'dir. Stopaj oranı %0'dır."),
    ("TLZ", "Faizsiz finans ilkelerine uyumlu şirketlerden seçilmiş olup, yalnızca Katılım Endeksi'nde yer alan şirketlerin ortaklık paylarından oluşmaktadır. Fon, Borsa İstanbul Katılım Endeksi'nde yer alan, uzun vadeli büyüme potansiyeli yüksek, sağlam finansal yapıya sahip şirketlerin hisse senedi paylarına yatırım yapmaktadır. Stopaj oranı %0'dır."),
    ("ANZ", "Fon portföyü, Türk bankaları başta olmak üzere, yurt dışı tahvil ihraç eden büyük şirketler ve TC Hazine'nin Eurobond'larından oluşur. Fonun portföyü, basiretli risk ile Dolar bazında en iyi getiriyi sağlamak için ortalama vade, faiz duyarlılığı (durasyon) ve de şirket riskleri açısından aktif olarak yönetilmektedir. Fon TL ve USD olarak fiyat açıklayıp aynı zamanda TL ve USD olarak işlem yapılabilmektedir. Fona nitelikli yatırımcılar yatırım yapabilmektedir. %17,5 stopaja tabidir."),
    ("PKP", "PKP Katılım Para Piyasası Fonu, faizsiz finans ilkelerine uygun olarak kısa vadeli ve düşük riskli yatırım araçlarına odaklanır. Fon portföyü, yüksek likidite ve sermaye korunumu hedefiyle, katılım esaslı para piyasası enstrümanlarından oluşur. Günlük nakit yönetimi ihtiyacı olan yatırımcılar için istikrarlı getiri ve kolay erişim imkânı sunar. Piyasa koşullarına hızlı uyum sağlayan portföy yapısı sayesinde, kısa vadeli dalgalanmalara karşı kontrollü bir risk profili sunmayı amaçlar. %17,5 stopaja tabidir. Fona nitelikli yatırımcılar yatırım yapabilmektedir."),
    ("AED", "Hisse senedi, altın, Eurobond ve döviz cinsinden enstrümanlara yatırım yaparak getiri sağlamayı hedeflemektedir. Varlık fiyatlarındaki aşırı yükselişler veya düşüşlerden faydalanmak üzere varlık sınıfları arasında ağırlık değişimleri (\"rebalancing\") yapılabilir. Fon minimum %51 hisse taşımaktadır. Portföyünde bir yıl tutan yatırımcılarına %0 stopaj avantajı sağlamaktadır."),
    ("AAS", "Ata Portföy Fon Sepeti Serbest Fon (AAS), yatırımcılara global çeşitlendirme stratejisi sunmaktadır. Fon, NYSE ve NASDAQ borsalarında yer alan, S&P 500 endeksinde bulunan şirketleri içeren ETF'lere yatırım yaparak güçlü bir portföy oluşturmayı amaçlamaktadır. Yatırımcılarına ABD ekonomisinin lider şirketlerine erişim imkânı sunarken, geniş sektör dağılımı ve sağlam bir temel ile uzun vadeli büyüme hedeflemektedir. Fon, yatırımlarını ETF'lerle (Borsa Yatırım Fonu) gerçekleştirmektedir. %80'in üzerinde yatırım fonlarına yatırım yapar."),
    ("YLC", "Artan nüfus, azalan kaynaklar ve gelişen teknolojiyle stratejik önemi her geçen gün artan tarım ve gıda sektörüne yatırım yaparak portföyünüzü çeşitlendirirken dünyanın geleceğine ortak olma fırsatı sağlar. Ata Portföy ve Nasdaq uzmanlığını kullanarak seçilen global tarım ve gıda sektörü hisse senetleri ve tarımsal emtialara yatırım yapmayı hedefler. Karşılaştırma ölçütü %65 NASDAQ Global Ex-Australia Agriculture Index (NQXAUAGR) + %35 BIST Gıda ve İçecek Getiri Endeksi (XGIDA)'dir. %17,5 stopaja tabidir."),
    ("RTG", "Ata Portföy ve NASDAQ uzmanlığını kullanarak global robotik teknolojiler sektörü hisse senetlerine yatırım yapmayı hedefler. Artan iş gücü maliyetleri ve teknolojik dönüşümün yarattığı fırsatları, yatırımcılara tek bir ürün halinde sunarak global ekonomik büyümeye uzun vadeli yatırım imkânı sağlar. Karşılaştırma ölçütü %65 NASDAQ CTA Artificial Intelligence & Robotics Index (NQROBO) + %35 BIST Bilişim Getiri Endeksi (XBLSM)'dir. Robotik Teknolojileri, Yapay Zeka, Otomasyon ve İnsansız Araçlar fonun yatırım yaptığı sektörlerdir. %17,5 stopaja tabidir."),
    ("JET", "Ata Portföy Havacılık ve Savunma Teknolojileri Değişken Fon; havacılık, uzay ve savunma alanlarında gelişen teknolojiler, artan stratejik işbirlikleri ve yükselen kamu harcamaları ile ortaya çıkan uzun vadeli küresel büyüme potansiyeline yatırım imkanı sunmayı hedefler. Fon, yerli ve yabancı havacılık, uzay ve savunma alanında faaliyet gösteren şirketlere yatırım yaparak bu sektörlerdeki teknolojik gelişmeler ve ekonomik büyümeye paydaş olma imkanı sunar. Türkiye, Amerika, Avrupa ve Asya şirketlerinden oluşturulan fon portföyü küresel çeşitlendirme imkanı sunmaktadır. Karşılaştırma ölçütü %65 Nasdaq US Benchmark Aerospace and Defense Total Return Index + %35 BIST Teknoloji Ağırlık Sınırlamalı Getiri Endeksi olarak belirlenmiştir. %17,5 stopaja tabidir."),
    ("URA", "Ata Portföy Enerji Değişken Fon öncelikli olarak nükleer enerji ile birlikte güneş, rüzgar, hidroelektrik, biyokütle gibi alternatif ve temiz enerji kaynaklarına yatırım yapmayı hedefler. Bu kapsamda: Global olarak uranyum madenciliği, nükleer santral teknolojisi ve inşaatı, ayrıca yeni nesil SMR (Küçük - Modüler Reaktör) geliştiren şirketlere yatırım yapılır. Yerli olarak biyogaz, jeotermal, güneş ve rüzgar enerjisi sektörlerinde faaliyet gösteren şirketlere yatırım yapılır. Fon, yüksek büyüme potansiyeline sahip şirketleri seçerek uzun vadede sürdürülebilir kazanç sağlamayı amaçlar. Karşılaştırma ölçütü %65 Solactive Global Uranium & Nuclear Components Total Return Index + %35 BIST Elektrik Getiri Endeksi olarak belirlenmiştir. %17,5 stopaja tabidir."),
    ("DGH", "Ata Portföy Para Piyasası Serbest (TL) Fon büyüklüğünün yaklaşık %60'ını mevduat gibi ürünlerde değerlendirirken, kalan kısmını portföy yöneticilerinin kriterlerine uygun özel sektör tahvil ve bonoları ile devlet tahviline yatıran bir fondur. Yatırım kararı, firmaların borçları ödeme gücüne dayanan detaylı kredi analizi sonucu alınır. Mevduat ile fonda sabit getiri hedeflenirken; özel sektör tahvil ve bonoları ile mevduat üzeri getiri hedeflenir. %17,5 stopaja tabidir."),
    ("PKF", "Ata Portföy Altın Katılım Fonu, altın ve altına dayalı faizsiz ürünlere yatırım yapar. Katılım esaslı olduğu için fon, faiz geliri içermemektedir. Altının piyasa koşullarına göre dalgalı fiyat hareketlerine rağmen uzun vadede enflasyona karşı yatırımlarınızı korumayı hedefler. %17,5 stopaja tabidir."),
]
for r, (kod, metin) in enumerate(fon_metinleri, start=2):
    ws.cell(row=r, column=1, value=kod).border = border
    c2 = ws.cell(row=r, column=2, value=metin)
    c2.border = border
    c2.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = 60
autosize(ws, [10, 110])
ws.freeze_panes = "A2"
ws.sheet_properties.tabColor = GREY_LIGHT

# ---------------------------------------------------------------
# 10. SABIT_METINLER
# ---------------------------------------------------------------
ws = wb.create_sheet("Sabit_Metinler")
ws["A1"] = "Anahtar"
ws["B1"] = "Metin"
style_header(ws, 1, 2)
sabit_metinler = [
    ("Kapak_Slogan", "Yatırım Yolculuğunuzda Yanınızdayız!"),
    ("Ata_Grubu_Tanitim", "Ata Holding'in kökleri, Atatürk Barajını yapan Ata İnşaat'a uzanır. Güneydoğu Anadolu Projesi (GAP) içinde yer alan Atatürk Barajı Hidroelektrik Santralini tamamlayan konsorsiyumun üç üyesinden biri olan Seri İnşaat, Grubumuz kurucu Sayın Ertuğrul Kurdoğlu'nun 1969 yılında kurduğu ilk mühendislik şirketidir.\n\nAta Grubu'nun faaliyet alanı zaman içinde finans, teknoloji, gıda, lojistik, gayrimenkul geliştirme, yapı malzemeleri, moda, turizm ve dış ticarete evrildi. Finans alanında ATA, gıda alanında TAB markasıyla önce çıkan Ata Holding, Türkiye başta olmak üzere pek çok ülkede Burger King master \"franchise\" sahibidir.\n\nDört kıta ve sekiz ülkede 30'dan fazla şirketle faaliyet gösteren Ata Holding'in yaygın portföy yapısı, risklere karşı koruma sağlamaktadır. Bu kapsamda Grubumuz 30.000'den fazla personele istihdam olanağı sunmaktadır."),
    ("Yatirim_100TL_Strateji_Baslik", "100 TL Yatırımınızı Nasıl Değerlendirmenizi Öneriyoruz?"),
    ("Yatirim_100TL_Strateji_Altbaslik", "Varlık dağılımı portföy yatırımında en önemli aşamadır."),
    ("Yatirim_100TL_Strateji_Giris", "100 TL'lik portföyü, 75 TL'si Türk Lirası, kalan 25 TL'si döviz bazında olacak şekilde aşağıdaki gibi dağıtıyoruz:"),
    # Sayfa 28 — Yatırım Fonları ile Portföy Yönetiminin Avantajları. Replaces
    # the earlier flat single-paragraph version (never actually rendered by
    # any page until now) — the real legacy page has two clearly separate
    # sections (a 6-item general list + a 3-item "müşteriye özel" list, see
    # Fon_Avantajlari_Maddeler below for the itemized bullets), confirmed by
    # rendering the actual PDF page rather than assuming the old blob was right.
    ("Fon_Yonetimi_Avantajlari_Giris", "Sermaye piyasası işlemlerini fon içinde yapmak kurumsal yatırımcılara önemli avantajlar sağlar."),
    ("Fon_Yonetimi_Avantajlari_Baslik2", "Müşteriye Özel Fon Tahsis Edilmesi;"),
    # Sayfa 29 — Yatırım Fonlarının Vergisel Avantajları. Also replaces an
    # earlier flat version — the real page has a left (bireysel) and right
    # (kurumlar + ANZ-specific) column, each with its own distinct pieces;
    # see Vergi_Bireysel_Fon_Listesi and Vergi_ANZ_Tablosu below for the
    # structured list/table parts.
    ("Vergi_Bireysel_Giris_1", "Fonlarda stopaj oranı  9 Temmuz 2025 itibarı ile %15'ten %17,5'a çıkartıldı."),
    ("Vergi_Bireysel_Giris_2", "%15 stopaja tabi olmayan fonlar:"),
    ("Vergi_Eurobond_Not", "Fon'dan elde edilen kazanç %17,5 stopaja tabi olup ayrıca beyana tabi değildir. (Örnek: ANZ)"),
    ("Vergi_Doviz_Not", "İbaresinde “Döviz” veya “Dolar” olan ve Dolar bazında fiyat açıklayan fonların %17,5 stopajı dolar kazanç üzerinden hesaplanır."),
    ("Vergi_Kurumlar_Not", "Yatırım fonlarından elde edilen gelirlerde stopaj oranı “0” olarak uygulanmaktadır ancak fon gelirleri kurumlar vergisi kapsamında beyan edilir ve kurumlar vergisine tabidir."),
    ("Vergi_ANZ_Tablo_Footnote", "* 2026'da yıllık 400.000 TL sermaye kazancı sınırı aşılırsa, faiz (kupon) ve itfada elde edilen faiz gelirinin tamamı beyan edilir. Alım satım kazancında maliyet ÜFE ile endekslenebilir."),
    # Sayfa 24 — F/K karşılaştırması. The first legacy bullet ("%26/%1
    # iskontolu") is computed live from Degerleme_CDS_MSCI's own F/K values
    # (BIST-100 vs Dünya, BIST-100 vs GOP) instead of stored as text here —
    # matches this project's established "compute, don't hand-type a number
    # that's already in the data" convention (e.g. PiyasaSenaryolariTablosu's
    # Değişim Potansiyeli). Only the second, non-numeric interpretive
    # sentence is stored as static text.
    ("Deger_FK_Yorum_2", "Türkiye ile ilgili olumsuz beklenti ve risklerin fiyatlandığı görülmektedir."),
    # Sayfa 25 — USD bazında BIST-100 endeksi yorumları (both purely
    # interpretive, no numbers to compute).
    ("Deger_USD_BIST_Yorum_1", "BIST100'ün Dolar bazında uzun vadede 150-500 aralığında dalgalanmıştır."),
    ("Deger_USD_BIST_Yorum_2", "150 seviyesine yakın yapılan hisse senedi yatırımları takip eden birkaç yılda Dolar bazında yüksek getiriler sağlamıştır."),
    # Sayfa 26 — uzun vadeli BIST getirileri yorumları.
    ("Bist_Yillik_Yorum_1", "Borsa İstanbul 1989 yılından beri hiç iki yıl üst üste eksi getirmedi."),
    ("Bist_Yillik_Yorum_2", "Hisse senetleri yüksek enflasyon dönemlerinde tarih boyunca tüm ülkelerde en yüksek getirileri sağlayan varlık sınıfı olmuştur."),
    # Sayfa 27 — MSCI Gelişmiş/Gelişmekte Olan Piyasalar.
    ("Msci_Yorum_1", "2010 yılından beri MSCI Gelişmekte Olan Piyasalar Endeksi (MSCI MXEF) yatay bir seyir izlerken, MSCI Gelişmiş Piyasalar Endeksi (MSCI MXWO) pozitif bir şekilde ayrışmıştır."),
    # Sayfa 30 — Ata Fonları katalog tablosu footnote (AED's risk-value
    # asterisk in Fon_Katalog.Stopaj, "%17,5*", refers to this note).
    ("Ata_Fonlari_Tablo_Footnote", "*Fon minimum %51 hisse taşımaktadır. Portföyünde bir yıl tutan yatırımcılarına %0 stopaj avantajı sağlamaktadır."),
    # Sayfa 31 — kapanış/iletişim sayfası.
    ("Iletisim_Baslik", "Varlık Yönetimi ve Yatırım Danışmanlığı Bölümü - İletişim"),
    ("Iletisim_Adres", "Ata Portföy Yönetimi A.Ş. Emirhan Cad. No:109 Atakule Kat:12 34349 Beşiktaş / İstanbul"),
    ("Iletisim_Telefon", "90 212 310 63 60"),
    ("Iletisim_Website", "www.ataportfoy.com.tr"),
    ("Yasal_Uyari", "Bu sunumda yer alan bilgiler Ata Portföy tarafından bilgilendirme amacıyla hazırlanmıştır. Yatırım bilgi, yorum ve tavsiyeleri yatırım danışmanlığı kapsamında değildir. Yatırım danışmanlığı hizmeti; aracı kurumlar, portföy yönetim şirketleri, mevduat kabul etmeyen bankalar ile müşteri arasında imzalanacak yatırım danışmanlığı sözleşmesi çerçevesinde sunulmaktadır. Burada yer alan yorum ve tavsiyeler, yorum ve tavsiyede bulunanların kişisel görüşlerine dayanmaktadır. Bu görüşlere ilişkin bilgiler güvenilirliğine inanılan kaynaklardan elde edilmiş olup, Ata Portföy bu bilgilerin doğruluğu hakkında garanti vermemektedir. Bu görüşler, mali durumunuz ile risk ve getiri tercihlerinize uygun olmayabilir. Bu nedenle, sadece burada yer alan bilgilere dayanılarak yatırım kararı verilmesi beklentilerinize uygun sonuçlar doğurmayabilir. Bu durumda, ortaya çıkabilecek sonuçlardan dolayı Ata Portföy sorumluluk kabul etmez. Ata Portföy, hiçbir şekil ve surette ön ihbara ve/veya ihtara gerek verebileceği herhangi bir zarardan sorumlu tutulamaz. Bu e-postanın virüs içermemesi için alınması gereken tedbirler alınmıştır. İşbu e-posta ve eklerinin kullanımından kaynaklı zarar veya kayıplardan sorumlu olmadığımızı kullanmadan önce virüs kontrol programlarınızı uygulamanızı tavsiye ederiz."),
]
for r, (key, metin) in enumerate(sabit_metinler, start=2):
    ws.cell(row=r, column=1, value=key).border = border
    c2 = ws.cell(row=r, column=2, value=metin)
    c2.border = border
    c2.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = 70
autosize(ws, [30, 110])
ws.freeze_panes = "A2"
ws.sheet_properties.tabColor = GREY_LIGHT

# ---------------------------------------------------------------
# 10b. FON_AVANTAJLARI_MADDELER — Sayfa 28's two bullet lists (Bolum
# distinguishes the "Genel" 6-item list from the "Musteriye_Ozel" 3-item one).
# ---------------------------------------------------------------
fon_avantajlari_data = [
    ("Genel", 1, "İşlemlerin fon içinde muhasebeleşmesi operasyonel ve vergisel açıdan kolaylık sağlar."),
    ("Genel", 2, "Nakit yönetimi açısından likidite kontrolü kolaydır. Nakit çıkışı valörü: T+1, T+2 veya T+3 olabilir."),
    ("Genel", 3, "Profesyonel yönetim ile yerli ve yabancı yatırım araçlarına erişim kolaylığı sağlar."),
    ("Genel", 4, "Yatırımcıların varlıkları, alınabilecek riskler göz önünde bulundurularak yüksek getiri hedefi ile profesyoneller tarafından yönetilir."),
    ("Genel", 5, "TEFAS üzerinden farklı varlık sınıflarını temsil eden fonlar seçilebilmektedir."),
    ("Genel", 6, "Fonun içerisindeki enstrümanların tek bir fiyattan alınıp satılabilmesi nedeniyle işlemlerin fon içerisinde yapılması müşterilere fırsat eşitliği sağlar."),
    ("Musteriye_Ozel", 1, "Sadece ilgili kuruma veya kişilere tahsis edilir."),
    ("Musteriye_Ozel", 2, "İçerik ve ürünler açısından yönetim avantajı sağlar."),
    ("Musteriye_Ozel", 3, "Fon büyüklüğüne bağlı olarak yönetim ücreti yönünden avantaj sağlar."),
]
build_tidy_sheet(
    "Fon_Avantajlari_Maddeler",
    ["Bolum", "Sira", "Metin"],
    [18, 8, 100],
    fon_avantajlari_data,
    tab_color=GREY_LIGHT,
    dropdown_cols={1: ["Genel", "Musteriye_Ozel"]},
    ornek_notu=False,
)

# ---------------------------------------------------------------
# 10c. VERGI_BIREYSEL_FON_LISTESI — Sayfa 29's left-column numbered list
# ("%15 stopaja tabi olmayan fonlar").
# ---------------------------------------------------------------
build_tidy_sheet(
    "Vergi_Bireysel_Fon_Listesi",
    ["Sira", "Aciklama"],
    [8, 100],
    [
        (1, "Hisse senedi yoğun fonları (Örnek: AAV, AYA ve TLZ)"),
        (2, "Bir yıldan daha fazla süreyle elde tutulan ve hisse oranı %51'den fazla olan fonlar (Örnek: AED)"),
    ],
    tab_color=GREY_LIGHT,
    ornek_notu=False,
)

# ---------------------------------------------------------------
# 10d. VERGI_ANZ_TABLOSU — Sayfa 29's right-column ANZ-specific tax table.
# Grup groups "1. Stopaj" and "2. Gelir Vergisi" under one merged
# "Vergi Oranı" row-header cell in the rendered table; Beyanname has none.
# ---------------------------------------------------------------
build_tidy_sheet(
    "Vergi_ANZ_Tablosu",
    ["Grup", "Satir", "Doviz_Mevduatinda", "Eurobond_Alim_Satim", "Eurobond_Fonu_Alirsa"],
    [14, 16, 26, 30, 34],
    [
        ("", "Beyanname", "Beyan edilmez", "Beyan edilir(*)", "Beyan edilmez"),
        ("Vergi Oranı", "1. Stopaj", "%25 Faiz vergisinden kesilir", "Yok", "%17,5 Katılımcı Fon'dan çıkarken Dolar değer kazancının üzerinden kesilir"),
        ("Vergi Oranı", "2. Gelir Vergisi", "Yok", "%15 - %40 Gelir Vergisi", "Yok"),
    ],
    tab_color=GREY_LIGHT,
    ornek_notu=False,
)

# ---------------------------------------------------------------
# 11. EKIP
# ---------------------------------------------------------------
ws = wb.create_sheet("Ekip")
headers = ["Ad_Soyad", "Unvan", "Telefon", "Eposta"]
for c, h in enumerate(headers, start=1):
    ws.cell(row=1, column=c, value=h)
style_header(ws, 1, len(headers))
ekip = [
    ("Etem Öztekin", "CEO, Yönetim Kurulu Üyesi", "0212 310 63 85", "etem.oztekin@ataportfoy.com.tr"),
    ("Şebnem Kaya", "Genel Müdür Yardımcısı", "0212 310 63 75", "skaya@ataportfoy.com.tr"),
    ("Funda Sevgi Tunal", "Direktör", "0212 310 60 21", "funda.tunal@ataportfoy.com.tr"),
    ("Bade Miray Süngü", "Müdür Yardımcısı", "0212 310 60 38", "bade.sungu@ataportfoy.com.tr"),
]
for r, row_data in enumerate(ekip, start=2):
    for c, val in enumerate(row_data, start=1):
        ws.cell(row=r, column=c, value=val).border = border
autosize(ws, [22, 30, 16, 32])
ws.freeze_panes = "A2"
ws.sheet_properties.tabColor = GREY_LIGHT

# ---------------------------------------------------------------
# 11b. EKIP_ORGANIZASYON — internal Fon Yönetimi / Fon Operasyon org chart
# (page "02" of the legacy deck), distinct from Ekip (the contact-page cards
# above): more people, grouped by department, no phone/email needed here.
# ---------------------------------------------------------------
ws = wb.create_sheet("Ekip_Organizasyon")
headers = ["Kart", "Bolum", "Ad_Soyad", "Unvan"]
for c, h in enumerate(headers, start=1):
    ws.cell(row=1, column=c, value=h)
style_header(ws, 1, len(headers))
# Bolum left blank = shown as card-top leadership (bold name + title beneath),
# not as a department section.
ekip_organizasyon = [
    ("Fon Yönetimi", "", "Etem Öztekin", "CEO, Board Member"),
    ("Fon Yönetimi", "Fon Yöneticileri", "Suna Tanrıverdi Fidan", ""),
    ("Fon Yönetimi", "Fon Yöneticileri", "Farshad Mirzazadeh, CFA", ""),
    ("Fon Yönetimi", "Fon Yöneticileri", "Samet Zağlı", ""),
    ("Fon Yönetimi", "Portföy Yöneticisi Yardımcıları", "Berk Tuna Subaşı", ""),
    ("Fon Yönetimi", "Portföy Yöneticisi Yardımcıları", "Elmas Öztürk", ""),
    ("Fon Yönetimi", "Araştırma", "Batuhan Demirci", ""),
    ("Fon Yönetimi", "Varlık Yönetimi ve Yatırım Danışmanlığı", "Şebnem Kaya", ""),
    ("Fon Yönetimi", "Varlık Yönetimi ve Yatırım Danışmanlığı", "Funda Sevgi Tunal", ""),
    ("Fon Yönetimi", "Varlık Yönetimi ve Yatırım Danışmanlığı", "Bade Miray Süngü", ""),
    ("Fon Yönetimi", "Varlık Yönetimi ve Yatırım Danışmanlığı", "Mete Tezel", ""),
    ("Fon Operasyon", "Fon Operasyon", "Niyazi Takmaz", ""),
    ("Fon Operasyon", "Fon Operasyon", "Güney Özütutar", ""),
    ("Fon Operasyon", "Fon Operasyon", "Aslı Serin", ""),
    ("Fon Operasyon", "Risk Yönetimi", "Seyhan Dönmezkuş", ""),
    ("Fon Operasyon", "İç Kontrol", "Sema Zorbacı", ""),
    ("Fon Operasyon", "Teftiş", "Gökhan Küçüoğlu", ""),
    ("Fon Operasyon", "İnsan Kaynakları", "Yaren Telatar", ""),
    ("Fon Operasyon", "Bilgi İşlem", "Kadir Mert Adnan", ""),
    ("Fon Operasyon", "Muhasebe", "Meltem Ayrım", ""),
]
for r, row_data in enumerate(ekip_organizasyon, start=2):
    for c, val in enumerate(row_data, start=1):
        ws.cell(row=r, column=c, value=val).border = border
autosize(ws, [16, 34, 24, 24])
ws.freeze_panes = "A2"
ws.sheet_properties.tabColor = GREY_LIGHT

# ---------------------------------------------------------------
# 11c. ATA_FONLARI_PERFORMANS_OZET — "Ata Fonları: Güçlü Performans" page
# (legacy page "03"): a curated ~23-row 1-year-return comparison across most
# of the ~24 funds Ata Portföy runs (broader than the 14 in Fon_Katalog,
# which only covers funds with their own individual page) plus the handful
# of benchmarks/macro series shown alongside them. Grup groups rows into the
# chart's 5 visual clusters (blank gap between each) — order matters, keep
# rows in display order, don't resort.
# ---------------------------------------------------------------
ws = wb.create_sheet("Ata_Fonlari_Performans_Ozet")
headers = ["Grup", "Ad", "Getiri_Yuzde", "Tip"]
for c, h in enumerate(headers, start=1):
    ws.cell(row=1, column=c, value=h)
style_header(ws, 1, len(headers))
fonlar_performans_ozet = [
    (1, "Ata İkinci Hisse Senedi Fonu", 18.3, "Fon"),
    (1, "Ata Kar Payı Ödeyen Hisse Senedi Fonu", 22.1, "Fon"),
    (1, "Ata Kar Payı Ödeyen Hisse Senedi Fonu (Temettü Dahil)", 34.7, "Fon"),
    (1, "Ata Katılım Hisse Senedi Fonu", 44.2, "Fon"),
    (1, "BIST 100 Getiri Endeksi", 31.3, "Benchmark"),
    (1, "BIST Temettü 25 Getiri Endeksi", 32.8, "Benchmark"),
    (2, "Ata Birinci Değişken Fon", 33.8, "Fon"),
    (3, "Ata Para Piyasası Serbest (TL) Fon", 50.6, "Fon"),
    (3, "Ata Para Piyasası Fonu", 46.4, "Fon"),
    (3, "KYD Mevduat Endeksi", 47.5, "Benchmark"),
    (4, "Ata Tarım ve Gıda Değişken Fon", 40.3, "Fon"),
    (4, "Ata Robotik Teknolojileri Değişken Fon", 60.4, "Fon"),
    (4, "Ata Havacılık ve Savunma Teknolojileri Değişken Fon", 26.7, "Fon"),
    (4, "Ata Enerji Değişken Fon", 38.3, "Fon"),
    (5, "Ata Altın Katılım Fonu", 37.4, "Fon"),
    (5, "Ata Fon Sepeti Serbest Fon", 33.9, "Fon"),
    (5, "Ata Dördüncü Serbest (Eurobond) Fon (TL Getiri)", 23.3, "Fon"),
    (5, "Ata Dördüncü Serbest (Eurobond) Fon (USD Getiri)", 5.4, "Fon"),
    (5, "USD/TL", 17.0, "Benchmark"),
    (5, "TÜFE (2026 - Yıllık)", 32.11, "Benchmark"),
]
for r, row_data in enumerate(fonlar_performans_ozet, start=2):
    for c, val in enumerate(row_data, start=1):
        ws.cell(row=r, column=c, value=val).border = border
autosize(ws, [8, 52, 14, 12])
ws.freeze_panes = "A2"
ws.sheet_properties.tabColor = AMBER_LIGHT

# ---------------------------------------------------------------
# 11e. ATA_FONLARI_PERFORMANS_OZET_YBB — the legacy deck's second "Güçlü
# Performans" page (page "04"), same idea as the one above but a
# Yılbaşından-Beri window instead of trailing-12-month, with its own
# (different, not just re-sorted) fund/benchmark selection and grouping.
# ---------------------------------------------------------------
ws = wb.create_sheet("Ata_Fonlari_Performans_Ozet_YBB")
headers = ["Grup", "Ad", "Getiri_Yuzde", "Tip"]
for c, h in enumerate(headers, start=1):
    ws.cell(row=1, column=c, value=h)
style_header(ws, 1, len(headers))
fonlar_performans_ozet_ybb = [
    (1, "İkinci Hisse Senedi Fonu", 15.8, "Fon"),
    (1, "Kar Payı Ödeyen Hisse Senedi Fonu", 15.2, "Fon"),
    (1, "Katılım Hisse Senedi (TL) Fonu", 34.2, "Fon"),
    (2, "Enerji Değişken Fon", 16.0, "Fon"),
    (2, "Havacılık ve Savunma Teknolojileri Fonu", 17.5, "Fon"),
    (2, "Tarım ve Gıda Değişken Fon", 29.2, "Fon"),
    (2, "Robotik Teknolojileri Değişken Fon", 37.7, "Fon"),
    (2, "BIST 100 Getiri Endeksi", 23.8, "Benchmark"),
    (2, "BIST Temettü 25 Getiri Endeksi", 26.4, "Benchmark"),
    (3, "ATA Birinci Değişken Fon", 26.4, "Fon"),
    (3, "KYD Mevduat Endeksi", 23.5, "Benchmark"),
    (4, "Ata Altın Katılım Fonu", 1.7, "Fon"),
    (4, "ATA 4. Serbest (EB) Fon (TL Getiri)", 10.8, "Fon"),
    (4, "ATA 4. Serbest (EB) Fon (USD Getiri)", 0.1, "Fon"),
    (4, "ATA Fon Sepeti Serbest Fon", 14.3, "Fon"),
    (4, "USD/TRY", 10.1, "Benchmark"),
    (5, "Ata Para Piyasası Serbest (TL) Fon", 25.1, "Fon"),
    (5, "Ata Para Piyasası Fonu", 21.9, "Fon"),
    (5, "TÜFE (2026 - Yılbaşından Beri)", 17.76, "Benchmark"),
]
for r, row_data in enumerate(fonlar_performans_ozet_ybb, start=2):
    for c, val in enumerate(row_data, start=1):
        ws.cell(row=r, column=c, value=val).border = border
autosize(ws, [8, 52, 14, 12])
ws.freeze_panes = "A2"
ws.sheet_properties.tabColor = AMBER_LIGHT

# ---------------------------------------------------------------
# 11f. YATIRIM_100TL_DAGILIM — the "100 TL Yatırımınızı Nasıl
# Değerlendirmenizi Öneriyoruz?" page's donut chart. Order matches the
# legacy page's own clockwise order starting from the top.
# ---------------------------------------------------------------
ws = wb.create_sheet("Yatirim_100TL_Dagilim")
headers = ["Varlik_Sinifi", "Yuzde"]
for c, h in enumerate(headers, start=1):
    ws.cell(row=1, column=c, value=h)
style_header(ws, 1, len(headers))
yatirim_100tl_dagilim = [
    ("Altın", 10),
    ("TL Sabit Getiri", 40),
    ("Türk Hisse Senedi", 35),
    ("Global Tematik Hisse Fonlar", 15),
]
for r, row_data in enumerate(yatirim_100tl_dagilim, start=2):
    for c, val in enumerate(row_data, start=1):
        ws.cell(row=r, column=c, value=val).border = border
autosize(ws, [30, 10])
ws.freeze_panes = "A2"
ws.sheet_properties.tabColor = GREY_LIGHT

# ---------------------------------------------------------------
# 11g. YATIRIM_100TL_MADDELER — the four sub-bullets under the intro line
# (Sabit_Metinler!Yatirim_100TL_Strateji_Giris), each ending in a bold TL
# amount. Kept structured (reason + amount as separate columns) rather than
# one hand-typed paragraph, on purpose — the earlier flat-string version was
# exactly the kind of single-blob text this project has otherwise avoided
# after the page-2 text-drift bug.
# ---------------------------------------------------------------
ws = wb.create_sheet("Yatirim_100TL_Maddeler")
headers = ["Sira", "Aciklama", "Tutar_TL"]
for c, h in enumerate(headers, start=1):
    ws.cell(row=1, column=c, value=h)
style_header(ws, 1, len(headers))
yatirim_100tl_maddeler = [
    (1, "Cazip değerleme oranları ve enflasyon üstü değer artışı potansiyeli nedeniyle hisse senetlerine", 35),
    (2, "Enflasyon ve cari açık beklentilerinin bozulmasının TL üzerinde yarattığı baskı nedeniyle altına", 10),
    (3, "Küresel piyasalarda tarım, gıda, teknoloji, havacılık ve savunma temalarına yatırım yapan tematik fonlara", 15),
    (4, "Cazip oranlardan dolayı kısa vadeli TL faize", 40),
]
for r, row_data in enumerate(yatirim_100tl_maddeler, start=2):
    for c, val in enumerate(row_data, start=1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.border = border
        if c == 2:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
autosize(ws, [8, 70, 12])
ws.freeze_panes = "A2"
ws.sheet_properties.tabColor = GREY_LIGHT

# ---------------------------------------------------------------
# 12. KAPAK_OZET
# ---------------------------------------------------------------
ws = wb.create_sheet("Kapak_Ozet")
headers = ["Tarih", "AUM_Milyar_TL", "Yatirim_Fon_Sayisi", "Serbest_Fon_Sayisi"]
for c, h in enumerate(headers, start=1):
    ws.cell(row=1, column=c, value=h)
style_header(ws, 1, len(headers))
ws.append(["2026-07-27", 25, 10, 14])
for c in range(1, 5):
    ws.cell(row=2, column=c).border = border
autosize(ws, [14, 16, 20, 20])
ws.freeze_panes = "A2"
ws.sheet_properties.tabColor = AMBER_LIGHT

# order sheets nicely
order = [
    "Talimatlar", "Kapak_Ozet", "Fon_Katalog", "Fon_Performans", "Fon_Performans_Ek_Kriter", "Fon_Fiyat_Serisi",
    "Fon_Portfoy_Dagilimi", "Fon_Getiri_Analizi_Detay", "Fon_Yillik_Getiri_Karsilastirma",
    "Fon_Temettu_Verimi",
    "Piyasa_Senaryolari", "Makro_Gostergeler", "Degerleme_CDS_MSCI", "Bist_Yillik_Getiriler",
    "Peer_PE_Karsilastirma", "Fon_Metinleri", "Sabit_Metinler",
    "Fon_Avantajlari_Maddeler", "Vergi_Bireysel_Fon_Listesi", "Vergi_ANZ_Tablosu",
    "Ekip", "Ekip_Organizasyon",
    "Ata_Fonlari_Performans_Ozet", "Ata_Fonlari_Performans_Ozet_YBB", "Yatirim_100TL_Dagilim",
    "Yatirim_100TL_Maddeler",
]
wb._sheets = [wb[name] for name in order]
wb.active = 0

wb.save("Veri_Kaynagi.xlsx")
print("OK, saved Veri_Kaynagi.xlsx with sheets:", wb.sheetnames)
