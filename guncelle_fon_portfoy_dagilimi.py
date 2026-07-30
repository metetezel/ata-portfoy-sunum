# Refreshes Fon_Portfoy_Dagilimi in Veri_Kaynagi.xlsx directly from TEFAS's
# own public API, instead of Mete hand-transcribing percentages off the
# legacy PDF. Fully automated, unlike Fon_Portfoy_Dagilimi's previous
# all-manual state — every fund this script covers gets a full replace on
# each run (no rows are ever hand-entered here, so there's nothing to
# merge-preserve, same pattern as Fon_Performans_Ek_Kriter).
#
# Source: https://www.tefas.gov.tr/api/funds/dagilimSiraliGetirT — the same
# endpoint TEFAS's own public "Fon Verileri" comparison page uses. Found by
# researching known open-source TEFAS API clients (pytefas), then verified
# directly against the live endpoint (2026-07-29) rather than trusted blind:
#   - The "fonKodu" request field is NOT actually a filter — the endpoint
#     always returns every registered fund (paginated via basSira/bitSira),
#     regardless of what's passed there. Filtering to our 14 codes happens
#     client-side after the fact.
#   - "basTarih"/"bitTarih" don't return day-by-day history either — the
#     endpoint always answers with each fund's single MOST RECENT
#     disclosure as of "bitTarih", so no "son gerçek tarih" trailing logic
#     (unlike every Price/Bench-sheet script in this project) is needed
#     here; one call always gets today's live snapshot directly.
#
# Chose to display TEFAS's own ~50 official regulatory categories as-is
# (Mete's explicit call, 2026-07-29) rather than force-mapping them into the
# legacy deck's old simplified 2-3-category groupings per fund — those
# categories don't correspond cleanly (e.g. TEFAS splits "Yatırım Fonu
# Katılma Payı"/"VİOP Teminatı"/"Ters Repo" separately where the legacy
# deck's own hand-made pie chart just lumped everything non-equity into a
# single "Nakit" slice) and building/maintaining a per-fund regrouping rule
# would be the "daha uzun sürecek" part of this task Mete anticipated.
# PortfoyPastasi.tsx already renders an arbitrary number of slices fine
# (its legend-height fix from the AED session scales with category count).
import datetime
import os

import openpyxl
import requests

VERI_KAYNAGI_YOLU = os.path.join(os.path.dirname(__file__), "Veri_Kaynagi.xlsx")

# Same 14-fund universe guncelle_fon_getirileri.py covers (includes AAL,
# which has a Fon_Katalog row but no deck page yet — kept in sync so this
# doesn't need its own separate fund list to maintain).
FON_KODLARI = {"AYA", "AAV", "TLZ", "URA", "JET", "YLC", "RTG", "AED", "AAS", "ANZ", "DGH", "AAL", "PKF", "PKP"}

# TEFAS's own short field codes -> Turkish display names. Kept as the full
# official list (not just the ones our 14 funds happen to use today) so a
# future fund needing a category none of the current ones use just works
# without touching this file. Verified against pytefas's own DIST_FIELDS
# English mapping, translated to the Turkish names TEFAS's own site uses.
KATEGORI_ADLARI = {
    "hs": "Yerli Hisse Senedi",
    "dt": "Devlet Tahvili",
    "hb": "Hazine Bonosu",
    "fb": "Finansman Bonosu",
    "ost": "Özel Sektör Tahvili",
    "bb": "Banka Bonosu",
    "vdm": "Varlığa Dayalı Menkul Kıymet",
    "eut": "Eurobond",
    "kibd": "Kamu Dış Borçlanma Araçları",
    "osdb": "Özel Sektör Dış Borçlanma Araçları",
    "kba": "Kamu İç Borçlanma Araçları (Döviz)",
    "dot": "Döviz Ödemeli Bono",
    "db": "Döviz Ödemeli Tahvil",
    "tpp": "Takasbank Para Piyasası",
    "bpp": "BİST Para Piyasası",
    "btaa": "BİST Vadeli İşlem (Alım)",
    "btas": "BİST Vadeli İşlem (Satım)",
    "r": "Repo",
    "tr": "Ters Repo",
    "vm": "Vadeli Mevduat",
    "vmtl": "Vadeli Mevduat (TL)",
    "vmd": "Vadeli Mevduat (Döviz)",
    "vmau": "Vadeli Mevduat (Altın)",
    "kh": "Katılma Hesabı",
    "khtl": "Katılma Hesabı (TL)",
    "khd": "Katılma Hesabı (Döviz)",
    "khau": "Katılma Hesabı (Altın)",
    "kks": "Kamu Kira Sertifikası",
    "kkstl": "Kamu Kira Sertifikası (TL)",
    "kksd": "Kamu Kira Sertifikası (Döviz)",
    "kksyd": "Kamu Dış Kira Sertifikası",
    "osks": "Özel Sektör Kira Sertifikası",
    "oksyd": "Özel Sektör Dış Kira Sertifikası",
    "km": "Kıymetli Maden",
    "kmbyf": "Kıymetli Maden BYF",
    "kmkba": "Kıymetli Maden (Kamu Borçlanma)",
    "kmkks": "Kıymetli Maden Kira Sertifikası",
    "ymk": "Yabancı Menkul Kıymet",
    "yba": "Yabancı Borçlanma Aracı",
    "ybkb": "Yabancı Kamu Borçlanma Aracı",
    "ybosb": "Yabancı Özel Sektör Borçlanma Aracı",
    "yhs": "Yabancı Hisse Senedi",
    "ybyf": "Yabancı BYF",
    "fkb": "Fon Katılma Belgesi",
    "yyf": "Yatırım Fonu Katılma Payı",
    "byf": "Borsa Yatırım Fonu",
    "gykb": "Gayrimenkul Yatırım Fonu",
    "gyy": "Gayrimenkul Yatırımı",
    "gsykb": "Girişim Sermayesi Yatırım Fonu",
    "gsyy": "Girişim Sermayesi Yatırımı",
    "t": "Türev Araçlar",
    "vint": "VİOP Teminatı",
    "gas": "Gayrimenkul Sertifikası",
    "d": "Diğer",
}


def tefas_dagilimlarini_cek():
    """Single POST to TEFAS's public breakdown endpoint, paginated wide
    enough (bitSira=5000) to comfortably cover every registered fund —
    confirmed 2026-07-29 that ~2000 already covers our whole 14-fund
    universe, 5000 leaves headroom for any future fund without ballooning
    the response unreasonably."""
    # basTarih/bitTarih must be real dates — the endpoint ignores the range
    # for what it returns (always each fund's single latest disclosure) but
    # errors on an empty/unparseable date string server-side.
    bugun = datetime.date.today().strftime("%Y%m%d")
    resp = requests.post(
        "https://www.tefas.gov.tr/api/funds/dagilimSiraliGetirT",
        headers={
            "Content-Type": "application/json",
            "Accept": "*/*",
            "Origin": "https://www.tefas.gov.tr",
            "Referer": "https://www.tefas.gov.tr/tr/fon-verileri",
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
        },
        json={"fonTipi": "YAT", "fonKodu": "", "basTarih": bugun, "bitTarih": bugun, "basSira": 1, "bitSira": 5000, "dil": "TR"},
        timeout=30,
    )
    resp.raise_for_status()
    data = resp.json()
    if data.get("errorMessage"):
        raise RuntimeError(f"TEFAS API hatası: {data['errorMessage']}")
    return data.get("resultList") or []


def main():
    tum_fonlar = tefas_dagilimlarini_cek()
    bizimkiler = {r["fonKodu"]: r for r in tum_fonlar if r.get("fonKodu") in FON_KODLARI}

    eksik = FON_KODLARI - bizimkiler.keys()
    if eksik:
        print(f"UYARI: TEFAS'ta bulunamayan fon kodları: {sorted(eksik)}")

    yeni_satirlar = []
    for fon_kodu, satir in bizimkiler.items():
        tarih = satir.get("tarih")
        for kisa_ad, yuzde in satir.items():
            if kisa_ad not in KATEGORI_ADLARI or yuzde is None:
                continue
            yeni_satirlar.append((tarih, fon_kodu, KATEGORI_ADLARI[kisa_ad], round(float(yuzde), 2)))

    wb = openpyxl.load_workbook(VERI_KAYNAGI_YOLU)
    ws = wb["Fon_Portfoy_Dagilimi"]
    ws.delete_rows(2, ws.max_row)
    for r, row_data in enumerate(yeni_satirlar, start=2):
        for c, val in enumerate(row_data, start=1):
            ws.cell(row=r, column=c, value=val)
    wb.save(VERI_KAYNAGI_YOLU)

    print(f"OK — TEFAS'tan {len(bizimkiler)}/{len(FON_KODLARI)} fon, {len(yeni_satirlar)} satır yazıldı.")


if __name__ == "__main__":
    main()
